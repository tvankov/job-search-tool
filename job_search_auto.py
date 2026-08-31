"""
job_search_auto.py
──────────────────
Headless version — runs without GUI, saves results to Excel.
Designed to be called by Windows Task Scheduler once a day.
"""

import os
import sys
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Resolve paths ─────────────────────────────────────────────────────────────
_BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH  = os.path.join(_BASE_DIR, "config.json")
_DELETED_PATH = os.path.join(_BASE_DIR, "deleted.json")
LOG_FILE      = os.path.join(_BASE_DIR, "job_search.log")
JOBS_DIR      = os.path.join(_BASE_DIR, "jobs")

# ── Make providers importable when run as a script ────────────────────────────
if _BASE_DIR not in sys.path:
    sys.path.insert(0, _BASE_DIR)

from providers import AuthError, build_provider_tasks, dedup_jobs

RESULTS_PER_PROVIDER = 50


# ── Config ────────────────────────────────────────────────────────────────────
def _load_deleted() -> set:
    """URLs the user permanently deleted in the GUI — never re-add these."""
    try:
        with open(_DELETED_PATH, encoding="utf-8") as f:
            return set(json.load(f))
    except Exception:
        return set()


def _load_config() -> dict:
    if os.path.exists(_CONFIG_PATH):
        try:
            with open(_CONFIG_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg: str):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# Route provider-level warnings (network / rate-limit / parse errors) to the
# same log file so a failing provider is visible, not silently swallowed.
_prov_log = logging.getLogger("jobsearch")
if not _prov_log.handlers:
    _prov_log.setLevel(logging.WARNING)
    _h = logging.FileHandler(LOG_FILE, encoding="utf-8")
    _h.setFormatter(logging.Formatter("[%(asctime)s] %(message)s", "%Y-%m-%d %H:%M:%S"))
    _prov_log.addHandler(_h)


# ── Fetch from all active providers ───────────────────────────────────────────
def fetch_jobs(what: str, where: str, country: str, cfg: dict) -> list:
    creds = {
        "adzuna_id":       cfg.get("app_id", ""),
        "adzuna_key":      cfg.get("app_key", ""),
        "reed_key":        cfg.get("reed_key", ""),
        "findwork_key":    cfg.get("findwork_key", ""),
        "jooble_key":      cfg.get("jooble_key", ""),
        "hh_token":        cfg.get("hh_token", ""),
        "themuse_key":     cfg.get("themuse_key", ""),
    }
    tasks = build_provider_tasks(
        use=cfg.get("providers", {}), creds=creds,
        what=what, where=where, country=country,
        results=RESULTS_PER_PROVIDER, sort_by="date")
    if not tasks:
        return []

    # Fetch all providers concurrently
    results_by_name = {}
    with ThreadPoolExecutor(max_workers=len(tasks)) as pool:
        futures = {pool.submit(fn): name for name, fn in tasks}
        for fut in as_completed(futures):
            name = futures[fut]
            try:
                jobs = fut.result()
                results_by_name[name] = jobs
                log(f"    {name}: {len(jobs)} jobs")
            except AuthError:
                results_by_name[name] = []
                log(f"    {name}: !! invalid API key — skipped")
            except Exception as e:
                results_by_name[name] = []
                log(f"    {name}: !! error — {e}")

    # Merge in task order so provider priority is preserved, dedup by URL
    return dedup_jobs([results_by_name.get(name, []) for name, _fn in tasks])


# ── Excel export ──────────────────────────────────────────────────────────────
def save_to_excel(jobs: list, job_title: str, search_info: dict) -> tuple:
    """Append new jobs to the persistent Excel file for this job title + location."""
    where = search_info.get("where", "").strip()
    label = f"{job_title} {where}".strip() if where else job_title
    safe  = label.replace(" ", "_")
    folder = os.path.join(JOBS_DIR, safe)
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"{safe}.xlsx")

    headers = ["Title", "Company", "Location",
               "Salary Min", "Salary Max", "Posted",
               "Link", "Description", "Source", "Imported On"]

    hdr_fill  = PatternFill("solid", start_color="0ea5e9")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    border    = Border(bottom=Side(style="thin", color="334155"))
    odd_fill  = PatternFill("solid", start_color="1e293b")
    even_fill = PatternFill("solid", start_color="162032")
    data_font = Font(name="Arial", size=10, color="F1F5F9")
    link_font = Font(name="Arial", size=10, color="38bdf8", underline="single")

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb["Job Results"]
        existing_urls = {str(row[6]) for row in ws.iter_rows(min_row=2, values_only=True) if row[6]}
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Results"
        existing_urls = set()
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = border
        ws.row_dimensions[1].height = 22

    next_row = ws.max_row + 1
    added = skipped = 0
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    deleted = _load_deleted()   # jobs the user removed in the GUI stay gone

    for job in jobs:
        url = job.get("url", "")
        if url in existing_urls or url in deleted:
            skipped += 1
            continue
        existing_urls.add(url)
        fill = odd_fill if (next_row % 2) else even_fill
        desc = str(job.get("description", "")).replace("\n", " ")[:300]

        row_data = [
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary_min") or "",
            job.get("salary_max") or "",
            str(job.get("created", ""))[:10],
            url,
            desc,
            job.get("source", ""),
            now_str,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 8))
            cell.font = (link_font if col == 7 and url else data_font)
            if col == 7 and url:
                cell.hyperlink = url
        ws.row_dimensions[next_row].height = 18
        next_row += 1
        added += 1

    for col, w in enumerate([40, 25, 22, 16, 16, 14, 40, 60, 18, 18], 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    # Search Info sheet
    if "Search Info" in wb.sheetnames:
        del wb["Search Info"]
    meta   = wb.create_sheet("Search Info")
    bg     = PatternFill("solid", start_color="1e293b")
    m_font = Font(name="Arial", size=10, color="F1F5F9")
    m_bold = Font(name="Arial", size=10, color="38bdf8", bold=True)
    total  = ws.max_row - 1

    for r, (lbl, val) in enumerate([
        ("Search query",       search_info.get("what", "")),
        ("Location",           search_info.get("where", "")),
        ("Last updated",       now_str),
        ("Total jobs",         total),
        ("Added this run",     added),
        ("Duplicates skipped", skipped),
    ], 1):
        c1 = meta.cell(row=r, column=1, value=lbl)
        c2 = meta.cell(row=r, column=2, value=val)
        c1.font = m_bold; c1.fill = bg
        c2.font = m_font; c2.fill = bg

    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 35
    meta.sheet_view.showGridLines = False
    ws.sheet_view.showGridLines   = False

    wb.save(path)
    return added, skipped, total


# ── Main ──────────────────────────────────────────────────────────────────────
def run():
    log("=" * 55)
    log("Job Search Auto-Run started")
    log("=" * 55)

    cfg     = _load_config()
    searches = cfg.get("searches", [])

    if not searches:
        log("!! No searches configured. Open the app → Auto Run tab → add a search.")
        log("=" * 55 + "\n")
        return

    active_providers = [k for k, v in cfg.get("providers", {}).items() if v]
    if not active_providers:
        log("!! No providers enabled. Open the app → chip bar → enable at least one.")
        log("=" * 55 + "\n")
        return

    log(f"Active providers: {', '.join(active_providers)}")
    total_added = 0

    for search in searches:
        what, where, country = search["what"], search["where"], search["country"]
        log(f"Searching: '{what}' in '{where}' ({country.upper()})")

        jobs = fetch_jobs(what, where, country, cfg)
        log(f"  >> {len(jobs)} total jobs fetched")

        if jobs:
            added, skipped, total = save_to_excel(jobs, what, search)
            log(f"  >> {added} new  |  {skipped} duplicates skipped  |  {total} total in file")
            total_added += added
        else:
            log("  >> No jobs returned")

    log(f"Done — {total_added} new jobs added across all searches")
    log("=" * 55 + "\n")


if __name__ == "__main__":
    run()
