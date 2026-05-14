import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import os
import sys
import json
import subprocess
import traceback
import threading
from datetime import datetime
from collections import Counter, defaultdict
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import io
try:
    from PIL import Image, ImageTk
    _PIL = True
except ImportError:
    _PIL = False
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from providers import (AuthError,
                       AdzunaProvider, ReedProvider, FindworkProvider,
                       ArbeitnowProvider, RemoteOKProvider,
                       JoobleProvider, TheMuseProvider,
                       BundesagenturProvider,
                       HeadHunterProvider, WeWorkRemotelyProvider,
                       RemotiveProvider, HimalayasProvider)

APP_VERSION = "1.0.1"

# ── Error logging ──────────────────────────────────────────────────────────────
def _log_error(exc_type, exc_value, exc_tb):
    try:
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "error.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n" + "=" * 70 + "\n")
            f.write(f"Date:    {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Version: {APP_VERSION}\n")
            f.write(f"OS:      {sys.platform}\n")
            f.write("Traceback:\n")
            f.writelines(traceback.format_exception(exc_type, exc_value, exc_tb))
    except Exception:
        pass
    sys.__excepthook__(exc_type, exc_value, exc_tb)

sys.excepthook = _log_error

# ── Config ────────────────────────────────────────────────────────────────────
def _app_dir() -> str:
    """Return the folder containing the exe (frozen) or the script (dev)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

_BASE_DIR    = _app_dir()
_CONFIG_PATH = os.path.join(_BASE_DIR, "config.json")

def _load_config():
    if os.path.exists(_CONFIG_PATH):
        try:
            with open(_CONFIG_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_config(data: dict):
    existing = _load_config()
    existing.update(data)
    with open(_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2)

_cfg    = _load_config()
APP_ID  = _cfg.get("app_id",  "")
APP_KEY = _cfg.get("app_key", "")

BG       = "#0f172a"
PANEL    = "#1e293b"
ACCENT   = "#0ea5e9"
ACCENT2  = "#38bdf8"
TEXT     = "#f1f5f9"
SUBTEXT  = "#94a3b8"
DANGER   = "#f87171"
SUCCESS  = "#34d399"
BORDER   = "#334155"
ROW_ODD  = "#1e293b"
ROW_EVEN = "#162032"

COUNTRIES = {
    "Algeria":              "dz",
    "Angola":               "ao",
    "Argentina":            "ar",
    "Armenia":              "am",
    "Australia":            "au",
    "Austria":              "at",
    "Azerbaijan":           "az",
    "Bahrain":              "bh",
    "Bangladesh":           "bd",
    "Belarus":              "by",
    "Belgium":              "be",
    "Bolivia":              "bo",
    "Brazil":               "br",
    "Bulgaria":             "bg",
    "Cambodia":             "kh",
    "Cameroon":             "cm",
    "Canada":               "ca",
    "Chile":                "cl",
    "China":                "cn",
    "Colombia":             "co",
    "Costa Rica":           "cr",
    "Croatia":              "hr",
    "Cyprus":               "cy",
    "Czech Republic":       "cz",
    "Denmark":              "dk",
    "Dominican Republic":   "do",
    "Ecuador":              "ec",
    "Egypt":                "eg",
    "El Salvador":          "sv",
    "Estonia":              "ee",
    "Ethiopia":             "et",
    "Finland":              "fi",
    "France":               "fr",
    "Georgia":              "ge",
    "Germany":              "de",
    "Ghana":                "gh",
    "Greece":               "gr",
    "Guatemala":            "gt",
    "Honduras":             "hn",
    "Hong Kong":            "hk",
    "Hungary":              "hu",
    "India":                "in",
    "Indonesia":            "id",
    "Ireland":              "ie",
    "Israel":               "il",
    "Italy":                "it",
    "Japan":                "jp",
    "Jordan":               "jo",
    "Kazakhstan":           "kz",
    "Kenya":                "ke",
    "Kuwait":               "kw",
    "Latvia":               "lv",
    "Lebanon":              "lb",
    "Lithuania":            "lt",
    "Luxembourg":           "lu",
    "Malaysia":             "my",
    "Malta":                "mt",
    "Mexico":               "mx",
    "Morocco":              "ma",
    "Mozambique":           "mz",
    "Myanmar":              "mm",
    "Netherlands":          "nl",
    "New Zealand":          "nz",
    "Nicaragua":            "ni",
    "Nigeria":              "ng",
    "Norway":               "no",
    "Oman":                 "om",
    "Pakistan":             "pk",
    "Panama":               "pa",
    "Paraguay":             "py",
    "Peru":                 "pe",
    "Philippines":          "ph",
    "Poland":               "pl",
    "Portugal":             "pt",
    "Qatar":                "qa",
    "Romania":              "ro",
    "Russia":               "ru",
    "Saudi Arabia":         "sa",
    "Senegal":              "sn",
    "Serbia":               "rs",
    "Singapore":            "sg",
    "Slovakia":             "sk",
    "Slovenia":             "si",
    "South Africa":         "za",
    "South Korea":          "kr",
    "Spain":                "es",
    "Sri Lanka":            "lk",
    "Sweden":               "se",
    "Switzerland":          "ch",
    "Taiwan":               "tw",
    "Tanzania":             "tz",
    "Thailand":             "th",
    "Tunisia":              "tn",
    "Turkey":               "tr",
    "UAE":                  "ae",
    "Uganda":               "ug",
    "UK":                   "gb",
    "Ukraine":              "ua",
    "Uruguay":              "uy",
    "USA":                  "us",
    "Venezuela":            "ve",
    "Vietnam":              "vn",
    "Zimbabwe":             "zw",
}

SORT_OPTIONS = {
    "Relevance": "relevance",
    "Date":      "date",
    "Salary ↑":  "salary_asc",
    "Salary ↓":  "salary_desc",
}

RESULTS_OPTIONS = [5, 10, 20, 50, 100, 200, 500]

# None = global (always supported)
PROVIDER_COUNTRIES = {
    "Adzuna":     {"gb","us","de","at","ch","fr","nl","be","pl","au","ca","br","in","nz","sg","za","ru","it","es"},
    "Reed":       {"gb"},
    "Findwork":   None,
    "Himalayas":  None,
    "Jooble":     None,

    "BA":         {"de"},
    "HeadHunter": {"ru","kz","by","uz","am","az","ge","kg","md","tj","tm"},
    "Arbeitnow":  {"de","at","ch","nl","be","fr","pl","it","es","gb","pt","se","no","dk","fi"},
    "The Muse":   {"us"},
    "RemoteOK":   None,
    "WWR":        None,
    "Remotive":   None,
}

# ── Main App ──────────────────────────────────────────────────────────────────
class JobSearchApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("J🔍B Search Tool  —  by Todor Vankov")
        self.geometry("980x680")
        self.configure(bg=BG)
        self.resizable(True, True)
        self.jobs = []
        try:
            import os
            if os.path.exists("icon.ico"):
                self.iconbitmap("icon.ico")
            elif os.path.exists("icon.png"):
                _ico = tk.PhotoImage(file="icon.png")
                self.iconphoto(True, _ico)
        except Exception:
            pass
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._style()
        self._header()
        self._update_banner = None
        self._footer()
        self._notebook()
        self.after(2000, self._check_for_updates)

    # ── Style ─────────────────────────────────────────────────────────────────
    def _style(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("TFrame",       background=BG)
        s.configure("TLabel",       background=BG, foreground=TEXT, font=("Segoe UI", 10))
        s.configure("TCombobox",
                    fieldbackground=PANEL, background=BORDER,
                    foreground=TEXT, selectbackground=PANEL,
                    selectforeground=TEXT,
                    insertcolor=TEXT,
                    bordercolor=BORDER, darkcolor=PANEL, lightcolor=PANEL,
                    arrowcolor=ACCENT2, arrowsize=12,
                    relief="flat", borderwidth=0,
                    font=("Segoe UI", 10), padding=(6, 4))
        s.map("TCombobox",
              fieldbackground=[("readonly", PANEL), ("disabled", PANEL), ("focus", PANEL)],
              background=[("active", BORDER), ("pressed", BORDER), ("!active", BORDER)],
              foreground=[("readonly", TEXT), ("disabled", SUBTEXT)],
              selectbackground=[("readonly", PANEL), ("focus", PANEL)],
              selectforeground=[("readonly", TEXT), ("focus", TEXT)],
              bordercolor=[("focus", ACCENT), ("!focus", BORDER)],
              arrowcolor=[("active", "white"), ("!active", ACCENT2)])
        self.option_add("*TCombobox*Listbox.background",   PANEL)
        self.option_add("*TCombobox*Listbox.foreground",   TEXT)
        self.option_add("*TCombobox*Listbox.selectBackground", ACCENT)
        self.option_add("*TCombobox*Listbox.selectForeground", "white")
        self.option_add("*TCombobox*Listbox.relief",       "flat")
        self.option_add("*TCombobox*Listbox.borderWidth",  "0")
        s.configure("Treeview",
                    background=PANEL, foreground=TEXT,
                    fieldbackground=PANEL, rowheight=28,
                    font=("Segoe UI", 10),
                    borderwidth=0, relief="flat",
                    bordercolor=PANEL)
        s.configure("Treeview.Heading",
                    background=ACCENT, foreground="white",
                    font=("Segoe UI", 10, "bold"),
                    relief="flat", borderwidth=0,
                    bordercolor=ACCENT)
        s.map("Treeview",
              background=[("selected", ACCENT), ("!selected", PANEL)],
              foreground=[("selected", "white")],
              fieldbackground=[("selected", PANEL), ("!selected", PANEL)])
        s.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
        for sb_style in ("Vertical.TScrollbar", "Horizontal.TScrollbar"):
            s.configure(sb_style,
                        background=BORDER,
                        troughcolor=BG,
                        bordercolor=BG,
                        arrowcolor=ACCENT2,
                        darkcolor=BORDER,
                        lightcolor=BORDER,
                        relief="flat",
                        borderwidth=0,
                        gripcount=0)
            s.map(sb_style,
                  background=[("active", ACCENT), ("pressed", ACCENT), ("!active", BORDER)],
                  arrowcolor=[("active", "white"), ("!active", ACCENT2)],
                  troughcolor=[("active", BG), ("!active", BG)])
        s.configure("TNotebook",        background=BG, borderwidth=0)
        s.configure("TNotebook.Tab",    background=PANEL, foreground=SUBTEXT,
                    font=("Segoe UI", 10), padding=(14, 6))
        s.map("TNotebook.Tab",
              background=[("selected", BG)],
              foreground=[("selected", ACCENT2)])
        # Tabless notebook — hide built-in tab bar completely
        s.configure("Tabless.TNotebook", tabmargins=[0, 0, 0, 0], borderwidth=0, padding=0)
        s.layout("Tabless.TNotebook", [("Notebook.client", {"sticky": "nswe"})])
        s.layout("Tabless.TNotebook.Tab", [])
        s.configure("Tabless.TNotebook.Tab",
                    padding=[0, 0, 0, 0], font=("Segoe UI", 1),
                    background=BG, foreground=BG, borderwidth=0)
        s.map("Tabless.TNotebook.Tab",
              background=[("selected", BG), ("active", BG)],
              foreground=[("selected", BG), ("active", BG)])

    # ── Shared logo widget ────────────────────────────────────────────────────
    def _make_logo(self, parent, letter_size, bg):
        frame = tk.Frame(parent, bg=bg)
        tk.Label(frame, text="J", bg=bg, fg=TEXT,
                 font=("Segoe UI", letter_size, "bold")).pack(side="left")
        tk.Label(frame, text="🔍", bg=bg, fg=ACCENT2,
                 font=("Segoe UI", int(letter_size * 0.72))).pack(
                     side="left", pady=(int(letter_size * 0.14), 0))
        tk.Label(frame, text="B", bg=bg, fg=TEXT,
                 font=("Segoe UI", letter_size, "bold")).pack(side="left")
        return frame

    # ── Header ────────────────────────────────────────────────────────────────
    def _header(self):
        bar = tk.Frame(self, bg=PANEL, height=54)
        bar.pack(fill="x")
        logo = self._make_logo(bar, 14, PANEL)
        logo.pack(side="left", padx=20, pady=10)
        self._btn(bar, "About", self._show_about,
                  color=PANEL, w=8).pack(side="right", padx=(0, 12), pady=10)
        self._btn(bar, "Help",  self._show_help,
                  color=PANEL, w=8).pack(side="right", padx=(0, 4),  pady=10)

    # ── Notebook ──────────────────────────────────────────────────────────────
    def _notebook(self):
        self._active_tab = 0
        _tab_names = ["Search", "Saved Results", "Auto Run", "Analytics", "Credentials"]

        # ── Custom tab bar ────────────────────────────────────────────────────
        tab_bar_wrap = tk.Frame(self, bg=BG)
        tab_bar_wrap.pack(fill="x")
        tk.Frame(tab_bar_wrap, bg=ACCENT, height=1).pack(fill="x", side="top")

        btn_row = tk.Frame(tab_bar_wrap, bg=BG)
        btn_row.pack(fill="x", side="top")

        self._tab_buttons    = []
        self._tab_indicators = []

        for i, name in enumerate(_tab_names):
            col = tk.Frame(btn_row, bg=BG)
            col.pack(side="left")
            btn = tk.Label(col, text=name, bg=BG,
                           fg=ACCENT if i == 0 else SUBTEXT,
                           font=("Segoe UI", 10, "bold" if i == 0 else "normal"),
                           padx=18, pady=9, cursor="hand2")
            btn.pack()
            ind = tk.Frame(col, height=2, bg=ACCENT if i == 0 else BG)
            ind.pack(fill="x")
            self._tab_buttons.append(btn)
            self._tab_indicators.append(ind)
            btn.bind("<Button-1>", lambda e, idx=i: self._switch_tab(idx))
            btn.bind("<Enter>",  lambda e, b=btn, idx=i: (
                b.config(fg=ACCENT2) if idx != self._active_tab else None))
            btn.bind("<Leave>",  lambda e, *_: self._update_tab_styles())

        tk.Frame(tab_bar_wrap, bg=ACCENT, height=1).pack(fill="x", side="bottom")

        # ── Plain frame stack — no ttk.Notebook border at all ────────────────
        self._nb_frame = tk.Frame(self, bg=BG)
        self._nb_frame.pack(fill="both", expand=True)

        self.tab_search      = tk.Frame(self._nb_frame, bg=BG)
        self.tab_saved       = tk.Frame(self._nb_frame, bg=BG)
        self.tab_autorun     = tk.Frame(self._nb_frame, bg=BG)
        self.tab_analytics   = tk.Frame(self._nb_frame, bg=BG)
        self.tab_credentials = tk.Frame(self._nb_frame, bg=BG)
        self._tab_frames = [self.tab_search, self.tab_saved, self.tab_autorun,
                            self.tab_analytics, self.tab_credentials]

        self._search_panel()
        self._results_panel()
        self._saved_panel()
        self._settings_panel()
        self._analytics_panel()
        self._credentials_panel()

        # Show first tab
        self.tab_search.pack(fill="both", expand=True)

    def _switch_tab(self, idx):
        if self._active_tab == idx:
            return
        self._tab_frames[self._active_tab].pack_forget()
        self._active_tab = idx
        self._tab_frames[idx].pack(fill="both", expand=True)
        self._update_tab_styles()
        self._on_tab_change(idx)

    def _update_tab_styles(self):
        for i, (btn, ind) in enumerate(zip(self._tab_buttons, self._tab_indicators)):
            if i == self._active_tab:
                btn.config(fg=ACCENT, font=("Segoe UI", 10, "bold"))
                ind.config(bg=ACCENT)
            else:
                btn.config(fg=SUBTEXT, font=("Segoe UI", 10))
                ind.config(bg=BG)

    # ── Search Panel ──────────────────────────────────────────────────────────
    def _search_panel(self):
        panel = tk.Frame(self.tab_search, bg=PANEL, pady=14)
        panel.pack(fill="x")

        row1 = tk.Frame(panel, bg=PANEL)
        row1.pack(fill="x", padx=20, pady=(0, 8))

        inp = tk.Frame(row1, bg=PANEL)
        inp.pack(side="left")

        tk.Label(inp, text="🔍  Job Title", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.what_var = tk.StringVar(value="Data Analyst")
        tk.Entry(inp, textvariable=self.what_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=22).grid(row=1, column=0, padx=(0, 12), ipady=6)

        tk.Label(inp, text="📍  Location", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=1, sticky="w")
        self.where_var = tk.StringVar(value="Berlin")
        tk.Entry(inp, textvariable=self.where_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=18).grid(row=1, column=1, padx=(0, 12), ipady=6)

        tk.Label(inp, text="🌍  Country", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=2, sticky="w")
        self.country_var = tk.StringVar(value="Germany")
        country_cb = ttk.Combobox(inp, textvariable=self.country_var, values=list(COUNTRIES.keys()),
                                  width=16, state="readonly")
        country_cb.grid(row=1, column=2, padx=(0, 12), ipady=4)
        self._bind_combobox_typeahead(country_cb, list(COUNTRIES.keys()))

        tk.Label(inp, text="Results", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=3, sticky="w")
        self.results_var = tk.IntVar(value=50)
        ttk.Combobox(inp, textvariable=self.results_var, values=RESULTS_OPTIONS,
                     width=6, state="readonly").grid(row=1, column=3, padx=(0, 12), ipady=4)


        btns = tk.Frame(row1, bg=PANEL)
        btns.pack(side="right", anchor="s")
        self._btn(btns, "🔍  Search", self._search, w=13).pack(side="left", padx=(0, 6))
        self._btn(btns, "💾  Save", self._export_excel, color="#0f766e", w=13).pack(side="left", padx=(0, 6))
        self._adv_btn_ref = self._btn(btns, "⚙  Filters ▾", lambda: self._toggle_advanced(), color="#334155", w=11)
        self._adv_btn_ref.pack(side="left")

        # Provider toggle chips — restore from config
        _pc = _load_config().get("providers", {})
        self.use_adzuna      = tk.BooleanVar(value=_pc.get("adzuna",     True))
        self.use_reed        = tk.BooleanVar(value=_pc.get("reed",       False))
        self.use_findwork    = tk.BooleanVar(value=_pc.get("findwork",   False))
        self.use_arbeitnow   = tk.BooleanVar(value=_pc.get("arbeitnow", False))
        self.use_remoteok    = tk.BooleanVar(value=_pc.get("remoteok",  False))
        self.use_jooble      = tk.BooleanVar(value=_pc.get("jooble",    False))
        self.use_themuse     = tk.BooleanVar(value=_pc.get("themuse",   False))
        self.use_bundesag    = tk.BooleanVar(value=_pc.get("bundesag",  False))
        self.use_headhunter  = tk.BooleanVar(value=_pc.get("headhunter",False))
        self.use_wwr         = tk.BooleanVar(value=_pc.get("wwr",       False))
        self.use_remotive    = tk.BooleanVar(value=_pc.get("remotive",  False))
        self.use_himalayas   = tk.BooleanVar(value=_pc.get("himalayas", False))

        providers_with_key = [
            ("Adzuna",         "19 countries",  self.use_adzuna),
            ("Reed",           "UK",            self.use_reed),
            ("Findwork",        "Tech/Remote",   self.use_findwork),
            ("Jooble",         "70+ countries", self.use_jooble),

            ("HeadHunter",     "RU/CIS",        self.use_headhunter),
        ]
        providers_no_key = [
            ("BA",             "DE",            self.use_bundesag, 20),
            ("Arbeitnow",      "EU",            self.use_arbeitnow),
            ("The Muse",       "USA",           self.use_themuse),
            ("RemoteOK",       "Remote",        self.use_remoteok),
            ("WWR",            "Remote",        self.use_wwr),
            ("Remotive",       "Remote",        self.use_remotive, 4),
            ("HIM",            "Remote",        self.use_himalayas),
        ]

        chip_row = tk.Frame(panel, bg=PANEL)

        cfg_now = _load_config()
        _key_ready = {
            "Adzuna":      bool(cfg_now.get("app_id") and cfg_now.get("app_key")),
            "Reed":        bool(cfg_now.get("reed_key")),
            "Findwork":    bool(cfg_now.get("findwork_key")),
            "Jooble":      bool(cfg_now.get("jooble_key")),

            "HeadHunter":  bool(cfg_now.get("hh_token")),
        }

        self._chip_info = []  # (label, var, set_supported_fn)

        def make_chip(parent, label, note, var, padx=8, ready=True):
            ON_BG    = ACCENT
            OFF_BG   = "#1e3a4a"
            DIS_BG   = "#111827"
            ON_NOTE  = "#bae6fd"
            OFF_NOTE = "#93c5b8"
            DIS_NOTE = "#374151"
            dot_color = SUCCESS if ready else "#64748b"
            _supported = [True]

            chip = tk.Frame(parent, bg=OFF_BG, padx=padx, pady=4, cursor="hand2")
            dot_lbl  = tk.Label(chip, text="⬤", bg=OFF_BG, fg=dot_color,
                                font=("Segoe UI", 6), cursor="hand2")
            name_lbl = tk.Label(chip, text=label, bg=OFF_BG,
                                font=("Segoe UI", 9, "bold"), cursor="hand2")
            note_lbl = tk.Label(chip, text=note, bg=OFF_BG, fg=OFF_NOTE,
                                font=("Segoe UI", 7), cursor="hand2")
            dot_lbl.pack(anchor="center")
            name_lbl.pack(anchor="center")
            note_lbl.pack(anchor="center")

            def refresh():
                if not _supported[0]:
                    return
                bg  = ON_BG  if var.get() else OFF_BG
                fg  = "white" if var.get() else SUBTEXT
                nfg = ON_NOTE if var.get() else OFF_NOTE
                chip.config(bg=bg, cursor="hand2")
                dot_lbl.config(bg=bg, cursor="hand2")
                name_lbl.config(bg=bg, fg=fg, cursor="hand2")
                note_lbl.config(bg=bg, fg=nfg, cursor="hand2")

            def toggle(_=None):
                if not _supported[0]:
                    return
                var.set(not var.get())
                refresh()
                self._save_provider_state()

            def on_enter(_=None):
                if not _supported[0]:
                    return
                bg = ACCENT2 if var.get() else "#2a4a5a"
                for w in (chip, dot_lbl, name_lbl, note_lbl):
                    w.config(bg=bg)

            def on_leave(_=None):
                refresh()

            def set_supported(supported: bool):
                _supported[0] = supported
                if not supported:
                    var.set(False)
                    for w in (chip, dot_lbl, name_lbl, note_lbl):
                        w.config(bg=DIS_BG, cursor="arrow")
                    name_lbl.config(fg=DIS_NOTE)
                    note_lbl.config(fg=DIS_NOTE)
                    dot_lbl.config(fg=dot_color)  # keep credential dot color
                else:
                    refresh()

            for w in (chip, dot_lbl, name_lbl, note_lbl):
                w.bind("<Button-1>", toggle)
                w.bind("<Enter>",    on_enter)
                w.bind("<Leave>",    on_leave)

            refresh()
            self._chip_info.append((label, var, set_supported))
            return chip

        # ── Advanced Filters (collapsible) ────────────────────────────────────
        self._adv_open = tk.BooleanVar(value=False)
        self._adv_frame = tk.Frame(panel, bg=PANEL)

        row2 = tk.Frame(self._adv_frame, bg=PANEL)
        row2.pack(fill="x", padx=20, pady=(6, 0))

        tk.Label(row2, text="Min. Salary (€)", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.salary_min_var = tk.StringVar()
        tk.Entry(row2, textvariable=self.salary_min_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 10),
                 width=12).grid(row=1, column=0, padx=(0, 12), ipady=4)

        tk.Label(row2, text="Max. Salary (€)", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=1, sticky="w")
        self.salary_max_var = tk.StringVar()
        tk.Entry(row2, textvariable=self.salary_max_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 10),
                 width=12).grid(row=1, column=1, padx=(0, 12), ipady=4)

        tk.Label(row2, text="Full-time only", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=2, sticky="w")
        self.fulltime_var = tk.BooleanVar()
        tk.Checkbutton(row2, variable=self.fulltime_var, bg=PANEL, fg=TEXT,
                       selectcolor=BG, activebackground=PANEL).grid(row=1, column=2, padx=(0, 12), sticky="w")

        tk.Label(row2, text="Permanent only", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=3, sticky="w")
        self.permanent_var = tk.BooleanVar()
        tk.Checkbutton(row2, variable=self.permanent_var, bg=PANEL, fg=TEXT,
                       selectcolor=BG, activebackground=PANEL).grid(row=1, column=3, padx=(0, 12), sticky="w")

        tk.Label(row2, text="Remote only", bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).grid(row=0, column=4, sticky="w")
        self.remote_var = tk.BooleanVar()
        tk.Checkbutton(row2, variable=self.remote_var, bg=PANEL, fg=TEXT,
                       selectcolor=BG, activebackground=PANEL).grid(row=1, column=4, sticky="w")

        prov_wrapper = tk.Frame(panel, bg=PANEL)
        prov_wrapper.pack(anchor="w", padx=20, pady=(10, 4))

        with_key_col = tk.Frame(prov_wrapper, bg=PANEL)
        with_key_col.pack(side="left")
        tk.Label(with_key_col, text="With key", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 3))
        chip_row_key = tk.Frame(with_key_col, bg=PANEL)
        chip_row_key.pack()
        for label, note, var in providers_with_key:
            make_chip(chip_row_key, label, note, var,
                      ready=_key_ready.get(label, False)).pack(side="left", padx=(0, 4))

        tk.Label(prov_wrapper, text="│", bg=PANEL, fg=BORDER,
                 font=("Segoe UI", 12)).pack(side="left", padx=8, anchor="s", pady=(0, 4))

        no_key_col = tk.Frame(prov_wrapper, bg=PANEL)
        no_key_col.pack(side="left")
        tk.Label(no_key_col, text="No key", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 3))
        chip_row_nokey = tk.Frame(no_key_col, bg=PANEL)
        chip_row_nokey.pack()
        for item in providers_no_key:
            label, note, var = item[0], item[1], item[2]
            extra = item[3] if len(item) > 3 else 8
            make_chip(chip_row_nokey, label, note, var, padx=extra, ready=True).pack(side="left", padx=(0, 4))

        tk.Label(panel, text="⚠  RemoteOK & WWR: job data accessed via public RSS feeds for personal use only.",
                 bg=PANEL, fg="#475569", font=("Segoe UI", 8)).pack(anchor="w", padx=20, pady=(4, 6))

        # update chips whenever country changes
        self.country_var.trace_add("write", lambda *_: self._update_chips_for_country())
        self._update_chips_for_country()

    def _update_chips_for_country(self):
        code = COUNTRIES.get(self.country_var.get(), "")
        for label, var, set_supported in self._chip_info:
            allowed = PROVIDER_COUNTRIES.get(label)
            supported = allowed is None or code in allowed
            set_supported(supported)

    def _toggle_advanced(self):
        if self._adv_open.get():
            self._adv_frame.pack_forget()
            self._adv_btn_ref.config(text="⚙  Filters ▾")
            self._adv_open.set(False)
        else:
            # pack after row1 (first child of panel)
            panel = self._adv_frame.master
            children = panel.pack_slaves()
            anchor = children[0] if children else None
            if anchor:
                self._adv_frame.pack(fill="x", after=anchor)
            else:
                self._adv_frame.pack(fill="x")
            self._adv_btn_ref.config(text="⚙  Filters ▲")
            self._adv_open.set(True)

    # ── Results Panel ─────────────────────────────────────────────────────────
    def _results_panel(self):
        frame = tk.Frame(self.tab_search, bg=BG, bd=0, highlightthickness=0)
        frame.pack(fill="both", expand=True, padx=16, pady=10)

        cols = ("title", "company", "location", "salary", "date", "source", "url")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse",
                                 style="Treeview")

        self._tree_col_labels = {
            "title":    "Job Title",
            "company":  "Company",
            "location": "Location",
            "salary":   "Salary",
            "date":     "Posted",
            "source":   "Source",
            "url":      "Link",
        }
        self._sort_col = None
        self._sort_rev = False

        for col, (label, width) in {
            "title":    ("Job Title", 280),
            "company":  ("Company",   150),
            "location": ("Location",  120),
            "salary":   ("Salary",     90),
            "date":     ("Posted",     85),
            "source":   ("Source",     80),
            "url":      ("Link",      110),
        }.items():
            self.tree.heading(col, text=label,
                              command=lambda c=col: self._sort_search_tree(c))
            self.tree.column(col, width=width, anchor="w")

        self.tree.tag_configure("odd",  background=ROW_ODD)
        self.tree.tag_configure("even", background=ROW_EVEN)

        sb_y = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        sb_x = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side="right",  fill="y")
        sb_x.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self._open_link)

        self._empty_state = tk.Frame(frame, bg=PANEL)
        self._empty_state.place(relx=0.5, rely=0.5, anchor="center")
        self._make_logo(self._empty_state, 28, PANEL).pack(pady=(0, 16))
        tk.Label(self._empty_state,
                 text="Enter a job title and click Search",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 13)).pack()
        tk.Label(self._empty_state,
                 text="Select one or more providers below",
                 bg=PANEL, fg="#475569", font=("Segoe UI", 10)).pack(pady=(4, 0))

    # ── Saved Results Panel ───────────────────────────────────────────────────
    def _saved_panel(self):
        saved_outer = tk.Frame(self.tab_saved, bg=BG, bd=0, highlightthickness=0)
        saved_outer.pack(fill="both", expand=True)

        # ── Left: file list ──────────────────────────────────────────────────
        left = tk.Frame(saved_outer, bg=PANEL, width=220, bd=0, highlightthickness=0)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
        tk.Frame(saved_outer, bg=BORDER, width=1).pack(side="left", fill="y")

        tk.Label(left, text="Saved Files", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(12, 6))

        tk.Frame(left, bg=BORDER, height=1).pack(fill="x", padx=8)

        list_frame = tk.Frame(left, bg=PANEL, bd=0, highlightthickness=0)
        list_frame.pack(fill="both", expand=True, padx=0, pady=(0, 8))

        sb = ttk.Scrollbar(list_frame, orient="vertical")
        sb.pack(side="right", fill="y")

        self.file_listbox = ttk.Treeview(
            list_frame, columns=("name",), show="",
            selectmode="browse", yscrollcommand=sb.set)
        self.file_listbox.column("name", width=190, anchor="w")
        self.file_listbox.pack(fill="both", expand=True)
        sb.config(command=self.file_listbox.yview)
        self.file_listbox.bind("<<TreeviewSelect>>", self._load_saved_file)

        btn_left = tk.Frame(left, bg=PANEL)
        btn_left.pack(pady=(0, 10))
        self._btn(btn_left, "Add to Auto Run", self._add_saved_to_autorun,
                  color="#0f766e", w=14).pack(side="left", padx=(8, 4))
        self._btn(btn_left, "Delete", self._delete_saved_file,
                  color="#7f1d1d", w=8).pack(side="left", padx=(0, 8))

        # ── Right: job table ─────────────────────────────────────────────────
        right = tk.Frame(saved_outer, bg=BG, bd=0, highlightthickness=0)
        right.pack(side="left", fill="both", expand=True)

        self.saved_info = tk.Label(right, text="Select a file to view its contents.",
                                   bg=BG, fg=SUBTEXT, font=("Segoe UI", 9))
        self.saved_info.pack(anchor="w", padx=12, pady=(10, 4))

        cols = ("title", "company", "location", "salary_min", "salary_max", "posted", "url")
        self.saved_tree = ttk.Treeview(right, columns=cols, show="headings", selectmode="browse",
                                       style="Treeview")

        self._saved_col_labels = {
            "title":      "Job Title",
            "company":    "Company",
            "location":   "Location",
            "salary_min": "Salary Min",
            "salary_max": "Salary Max",
            "posted":     "Posted",
            "url":        "Link",
        }
        self._saved_sort = {"col": None, "rev": False}

        for col, (label, width) in {
            "title":      ("Job Title",   300),
            "company":    ("Company",     150),
            "location":   ("Location",    120),
            "salary_min": ("Salary Min",   90),
            "salary_max": ("Salary Max",   90),
            "posted":     ("Posted",       90),
            "url":        ("Link",        120),
        }.items():
            self.saved_tree.heading(col, text=label,
                command=lambda c=col: self._sort_treeview(
                    self.saved_tree, c, self._saved_col_labels, self._saved_sort))
            self.saved_tree.column(col, width=width, anchor="w")

        self.saved_tree.tag_configure("odd",  background=ROW_ODD)
        self.saved_tree.tag_configure("even", background=ROW_EVEN)

        sb_y = ttk.Scrollbar(right, orient="vertical",   command=self.saved_tree.yview)
        sb_x = ttk.Scrollbar(right, orient="horizontal", command=self.saved_tree.xview)
        self.saved_tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side="right",  fill="y")
        sb_x.pack(side="bottom", fill="x")
        self.saved_tree.pack(fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        self.saved_tree.bind("<Double-1>", self._open_saved_link)

        self._saved_files = {}  # display_name -> full_path
        self._refresh_saved_list()

    def _refresh_saved_list(self):
        self._saved_files.clear()
        for row in self.file_listbox.get_children():
            self.file_listbox.delete(row)
        jobs_dir = os.path.join(_app_dir(), "jobs")
        if not os.path.isdir(jobs_dir):
            return
        for root, _, files in os.walk(jobs_dir):
            for f in files:
                if f.endswith(".xlsx"):
                    full = os.path.join(root, f)
                    label = os.path.splitext(f)[0].replace("_", " ")
                    self._saved_files[label] = full
        for i, label in enumerate(sorted(self._saved_files)):
            self.file_listbox.insert("", "end", iid=label,
                                     values=(label,),
                                     tags=("odd" if i % 2 else "even",))
        self._analytics_populate_files()

    def _load_saved_file(self, event=None):
        sel = self.file_listbox.selection()
        if not sel:
            return
        label = self.file_listbox.item(sel[0])["values"][0]
        path  = self._saved_files.get(label)
        if not path or not os.path.exists(path):
            return

        self._saved_sort = {"col": None, "rev": False}
        for c, lbl in self._saved_col_labels.items():
            self.saved_tree.heading(c, text=lbl)
        for row in self.saved_tree.get_children():
            self.saved_tree.delete(row)

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if "Job Results" not in wb.sheetnames:
                self.saved_info.config(text="No 'Job Results' sheet found in this file.")
                return
            ws = wb["Job Results"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for i, r in enumerate(rows):
                title, company, location, sal_min, sal_max, posted, url = (
                    (r[0] or ""), (r[1] or ""), (r[2] or ""),
                    (r[3] or ""), (r[4] or ""), (r[5] or ""), (r[6] or ""))
                self.saved_tree.insert("", "end", iid=str(i),
                                       tags=("odd" if i % 2 else "even",),
                                       values=(title, company, location,
                                               sal_min, sal_max, posted, url))
            wb.close()
            self.saved_info.config(
                text=f"{label}  —  {len(rows)} jobs  |  {path}", fg=SUBTEXT)
        except Exception as e:
            self.saved_info.config(text=f"Error reading file: {e}", fg=DANGER)

    def _add_saved_to_autorun(self):
        sel = self.file_listbox.selection()
        if not sel:
            return
        label = self.file_listbox.item(sel[0])["values"][0]
        path  = self._saved_files.get(label)
        if not path:
            return
        file_slug = os.path.splitext(os.path.basename(path))[0]
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        already  = any(
            (s.get("what", "").strip() + "_" + s.get("where", "").strip())
            .replace(" ", "_").lower() == file_slug.lower()
            for s in searches
        )
        if already:
            messagebox.showinfo("Auto Run", f'"{label}" is already in Auto Run.')
            return
        what = where = country = ""
        try:
            wb_r = load_workbook(path, read_only=True, data_only=True)
            if "Search Info" in wb_r.sheetnames:
                for row in wb_r["Search Info"].iter_rows(values_only=True):
                    if row[0] == "Search query":
                        what = str(row[1] or "")
                    elif row[0] == "Location":
                        where = str(row[1] or "")
                    elif row[0] == "Country":
                        country = COUNTRIES.get(str(row[1] or ""), "de")
            wb_r.close()
        except Exception:
            pass
        if not what:
            what = file_slug.replace("_", " ")
        searches.append({"what": what, "where": where, "country": country or "de"})
        _save_config({"searches": searches})
        self._autorun_refresh()
        messagebox.showinfo("Auto Run", f'"{label}" added to Auto Run.')

    def _delete_saved_file(self):
        sel = self.file_listbox.selection()
        if not sel:
            return
        label = self.file_listbox.item(sel[0])["values"][0]
        path  = self._saved_files.get(label)
        if not path or not os.path.exists(path):
            return
        if not messagebox.askyesno("Delete File",
                                   f"Delete '{label}'?\n\n{path}\n\nThis cannot be undone."):
            return
        try:
            os.remove(path)
            folder = os.path.dirname(path)
            if not os.listdir(folder):
                os.rmdir(folder)
        except Exception as e:
            messagebox.showerror("Error", f"Could not delete file:\n{e}")
            return
        for row in self.saved_tree.get_children():
            self.saved_tree.delete(row)
        self.saved_info.config(text="Select a file to view its contents.", fg=SUBTEXT)
        self._refresh_saved_list()

        file_slug = os.path.splitext(os.path.basename(path))[0].lower()
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        filtered = [s for s in searches
                    if (s.get("what", "").strip() + "_" + s.get("where", "").strip())
                       .replace(" ", "_").lower() != file_slug]
        _save_config({"searches": filtered})
        self._autorun_refresh()

    def _open_saved_link(self, event=None):
        sel = self.saved_tree.selection()
        if sel:
            url = self.saved_tree.item(sel[0])["values"][6]
            if url:
                __import__("webbrowser").open(str(url))

    # ── Auto Run Panel ────────────────────────────────────────────────────────
    def _settings_panel(self):
        outer = tk.Frame(self.tab_autorun, bg=BG, bd=0, highlightthickness=0)
        outer.pack(fill="both", expand=True, padx=30, pady=24)

        # ── Schedule section ─────────────────────────────────────────────────
        section = tk.Frame(outer, bg=PANEL, pady=18)
        section.pack(fill="x")

        sec_left = tk.Frame(section, bg=PANEL)
        sec_left.pack(side="left", fill="both", expand=True)

        title_row = tk.Frame(sec_left, bg=PANEL)
        title_row.pack(anchor="w", padx=20, pady=(0, 4))
        tk.Label(title_row, text="Scheduled Auto-Run", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(side="left")
        self.sched_active_lbl = tk.Label(title_row, text="⬤  Checking…", bg=PANEL,
                                         fg=SUBTEXT, font=("Segoe UI", 9))
        self.sched_active_lbl.pack(side="left", padx=(14, 0), pady=(2, 0))
        tk.Label(sec_left, text="Run all searches below automatically every day at a set time.",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 14))

        time_row = tk.Frame(sec_left, bg=PANEL)
        time_row.pack(anchor="w", padx=20, pady=(0, 14))
        tk.Label(time_row, text="Time (HH:MM)", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.sched_time_var = tk.StringVar(value=_load_config().get("sched_time", "08:00"))
        tk.Entry(time_row, textvariable=self.sched_time_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=10).grid(row=1, column=0, ipady=6, padx=(0, 16))

        btn_row = tk.Frame(sec_left, bg=PANEL)
        btn_row.pack(anchor="w", padx=20)
        self._btn(btn_row, "Schedule Daily", self._schedule, w=16).pack(side="left", padx=(0, 10))
        self._btn(btn_row, "Remove Schedule", self._unschedule, color="#7f1d1d", w=16).pack(side="left")

        self.sched_status = tk.Label(sec_left, text="", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9))
        self.sched_status.pack(anchor="w", padx=20, pady=(10, 0))

        # ── Right info panel ──────────────────────────────────────────────────
        sec_right = tk.Frame(section, bg=PANEL, width=220)
        sec_right.pack(side="right", fill="y", padx=(0, 0), pady=0)
        sec_right.pack_propagate(False)

        tk.Frame(sec_right, bg=BORDER, width=1).pack(side="left", fill="y")

        inner = tk.Frame(sec_right, bg=PANEL)
        inner.place(relx=0.5, rely=0.5, anchor="center")

        tk.Label(inner, text="Last Auto-Run", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 8, "bold")).pack(anchor="center")

        tk.Frame(inner, bg=BORDER, height=1).pack(fill="x", pady=6)

        self.log_date_lbl = tk.Label(inner, text="—", bg=PANEL, fg=ACCENT2,
                                     font=("Segoe UI", 11, "bold"), cursor="hand2")
        self.log_date_lbl.pack(anchor="center")
        self.log_date_lbl.bind("<Button-1>", lambda e: self._show_run_history())

        self.log_searches_lbl = tk.Label(inner, text="", bg=PANEL, fg=SUBTEXT,
                                         font=("Segoe UI", 9))
        self.log_searches_lbl.pack(anchor="center", pady=(6, 0))

        self.log_jobs_lbl = tk.Label(inner, text="", bg=PANEL, fg=SUCCESS,
                                     font=("Segoe UI", 10, "bold"))
        self.log_jobs_lbl.pack(anchor="center", pady=(2, 0))

        self._refresh_sched_status()
        self._refresh_log_panel()

        # ── Searches list ─────────────────────────────────────────────────────
        tk.Frame(outer, bg=BORDER, height=1).pack(fill="x", pady=(18, 0))

        searches_hdr = tk.Frame(outer, bg=BG)
        searches_hdr.pack(fill="x", pady=(12, 6))
        tk.Label(searches_hdr, text="Searches to run", bg=BG, fg=ACCENT2,
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        self._btn(searches_hdr, "Remove Selected", self._autorun_remove,
                  color="#7f1d1d", w=16).pack(side="right", padx=(6, 0))
        self._btn(searches_hdr, "Add from Search Tab",
                  self._autorun_add_current, color="#334155", w=20).pack(side="right")

        # Treeview
        s_cols = ("what", "where", "country", "file")
        self.runs_tree = ttk.Treeview(outer, columns=s_cols, show="headings",
                                      selectmode="browse", height=5, style="Treeview")
        self._runs_col_labels = {
            "what":    "Job Title",
            "where":   "Location",
            "country": "Country",
            "file":    "Saves to",
        }
        self._runs_sort = {"col": None, "rev": False}
        for col, (label, width) in {
            "what":    ("Job Title", 200),
            "where":   ("Location",  140),
            "country": ("Country",   120),
            "file":    ("Saves to",  260),
        }.items():
            self.runs_tree.heading(col, text=label,
                command=lambda c=col: self._sort_treeview(
                    self.runs_tree, c, self._runs_col_labels, self._runs_sort))
            self.runs_tree.column(col, width=width, anchor="w")
        self.runs_tree.tag_configure("odd",  background=ROW_ODD)
        self.runs_tree.tag_configure("even", background=ROW_EVEN)

        runs_sb = ttk.Scrollbar(outer, orient="vertical", command=self.runs_tree.yview)
        self.runs_tree.configure(yscrollcommand=runs_sb.set)
        runs_sb.pack(side="right", fill="y")
        self.runs_tree.pack(fill="x", pady=(0, 6))

        # Add form + Remove button
        add_row = tk.Frame(outer, bg=BG)
        add_row.pack(fill="x", pady=(0, 4))

        tk.Label(add_row, text="Job Title", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 8)).grid(row=0, column=0, sticky="w")
        tk.Label(add_row, text="Location", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 8)).grid(row=0, column=1, sticky="w", padx=(8, 0))
        tk.Label(add_row, text="Country", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 8)).grid(row=0, column=2, sticky="w", padx=(8, 0))

        self.run_what_var  = tk.StringVar()
        self.run_where_var = tk.StringVar()
        self.run_country_var = tk.StringVar(value="Germany")

        tk.Entry(add_row, textvariable=self.run_what_var, bg=PANEL, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 10),
                 width=22).grid(row=1, column=0, ipady=5)
        tk.Entry(add_row, textvariable=self.run_where_var, bg=PANEL, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 10),
                 width=18).grid(row=1, column=1, ipady=5, padx=(8, 0))
        ttk.Combobox(add_row, textvariable=self.run_country_var,
                     values=list(COUNTRIES.keys()), width=14,
                     state="readonly").grid(row=1, column=2, ipady=3, padx=(8, 0))
        self._btn(add_row, "Add", self._autorun_add, w=6
                  ).grid(row=1, column=3, padx=(10, 0))

        self._autorun_refresh()

        # ── Search History ────────────────────────────────────────────────────
        tk.Frame(outer, bg=BORDER, height=1).pack(fill="x", pady=(16, 0))

        hist_header = tk.Frame(outer, bg=BG)
        hist_header.pack(fill="x", pady=(12, 6))
        tk.Label(hist_header, text="Search History  (last 30)", bg=BG, fg=ACCENT2,
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        self._btn(hist_header, "Refresh", self._refresh_history,
                  color="#334155", w=10).pack(side="right")

        hist_cols = ("datetime", "query", "location", "fetched", "new", "skipped")
        self.hist_tree = ttk.Treeview(outer, columns=hist_cols, show="headings",
                                      selectmode="none", height=8, style="Treeview")
        for col, (label, width) in {
            "datetime": ("Date / Time",  140),
            "query":    ("Search",       160),
            "location": ("Location",     110),
            "fetched":  ("Fetched",       70),
            "new":      ("New",           60),
            "skipped":  ("Skipped",       70),
        }.items():
            self.hist_tree.heading(col, text=label)
            self.hist_tree.column(col, width=width, anchor="w")

        self.hist_tree.tag_configure("odd",  background=ROW_ODD)
        self.hist_tree.tag_configure("even", background=ROW_EVEN)

        hist_sb = ttk.Scrollbar(outer, orient="vertical", command=self.hist_tree.yview)
        self.hist_tree.configure(yscrollcommand=hist_sb.set)
        hist_sb.pack(side="right", fill="y")
        self.hist_tree.pack(fill="both", expand=True, pady=(0, 10))

        self._refresh_history()

    # ── Auto Run helpers ──────────────────────────────────────────────────────
    def _refresh_sched_status(self):
        try:
            result = subprocess.run(
                ["schtasks", "/query", "/tn", "JobSearchDaily", "/fo", "LIST"],
                capture_output=True, text=True)
            if result.returncode == 0:
                time_str = _load_config().get("sched_time", "")
                label = f"⬤  Active{f'  {time_str}' if time_str else ''}"
                self.sched_active_lbl.config(text=label, fg=SUCCESS)
            else:
                self.sched_active_lbl.config(text="⬤  Not scheduled", fg=DANGER)
        except Exception:
            self.sched_active_lbl.config(text="⬤  Unknown", fg=SUBTEXT)

    def _refresh_log_panel(self):
        log_path = os.path.join(_app_dir(), "job_search.log")
        if not os.path.exists(log_path):
            self.log_date_lbl.config(text="No log yet.")
            self.log_searches_lbl.config(text="")
            self.log_jobs_lbl.config(text="")
            return
        try:
            with open(log_path, encoding="utf-8") as f:
                lines = [l.rstrip() for l in f if l.strip()]
            # find last run block
            start_idx = None
            for i in range(len(lines) - 1, -1, -1):
                if "Auto-Run started" in lines[i]:
                    start_idx = i
                    break
            if start_idx is None:
                self.log_date_lbl.config(text="No runs yet.")
                self.log_searches_lbl.config(text="")
                self.log_jobs_lbl.config(text="")
                return
            block = lines[start_idx:]
            ts       = block[0][1:17] if block[0].startswith("[") else "?"
            searches = sum(1 for l in block if "Searching:" in l)
            new_jobs = 0
            for l in block:
                if "Done —" in l and "new jobs" in l:
                    try:
                        new_jobs = int(l.split("Done —")[1].split("new")[0].strip())
                    except Exception:
                        pass
            self.log_date_lbl.config(text=ts)
            self.log_searches_lbl.config(text=f"{searches} search{'es' if searches != 1 else ''}")
            self.log_jobs_lbl.config(text=f"+{new_jobs} new jobs")
        except Exception:
            self.log_date_lbl.config(text="Could not read log.")
            self.log_searches_lbl.config(text="")
            self.log_jobs_lbl.config(text="")

    def _show_run_history(self):
        log_path = os.path.join(_app_dir(), "job_search.log")
        runs = []
        try:
            with open(log_path, encoding="utf-8") as f:
                lines = [l.rstrip() for l in f if l.strip()]
            block = []
            for line in lines:
                if "Auto-Run started" in line:
                    block = [line]
                elif block:
                    block.append(line)
                    if "Done —" in line:
                        ts       = block[0][1:17] if block[0].startswith("[") else "?"
                        searches = sum(1 for l in block if "Searching:" in l)
                        new_jobs = 0
                        try:
                            new_jobs = int(line.split("Done —")[1].split("new")[0].strip())
                        except Exception:
                            pass
                        runs.append((ts, searches, new_jobs))
                        block = []
        except Exception:
            pass

        win = tk.Toplevel(self, bg=BG)
        win.title("Auto-Run History")
        win.geometry("520x420")
        win.resizable(False, False)

        tk.Label(win, text="Auto-Run History", bg=BG, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(16, 8))

        cols = ("date", "searches", "new_jobs")
        tree = ttk.Treeview(win, columns=cols, show="headings", selectmode="none",
                            style="Treeview")
        for col, label, width in [("date", "Date / Time", 200),
                                   ("searches", "Searches", 100),
                                   ("new_jobs", "New Jobs", 100)]:
            tree.heading(col, text=label)
            tree.column(col, width=width, anchor="w")
        tree.tag_configure("odd",  background=ROW_ODD)
        tree.tag_configure("even", background=ROW_EVEN)

        sb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y", padx=(0, 10), pady=(0, 10))
        tree.pack(fill="both", expand=True, padx=(20, 0), pady=(0, 10))

        for i, (ts, searches, new_jobs) in enumerate(reversed(runs[-100:])):
            tree.insert("", "end", tags=("odd" if i % 2 else "even",),
                        values=(ts, f"{searches} search{'es' if searches != 1 else ''}", f"+{new_jobs}"))

        if not runs:
            tree.insert("", "end", values=("No runs logged yet.", "", ""))

    def _autorun_refresh(self):
        for row in self.runs_tree.get_children():
            self.runs_tree.delete(row)
        searches = _load_config().get("searches", [])
        country_names = {v: k for k, v in COUNTRIES.items()}
        for i, s in enumerate(searches):
            what         = s.get("what", "")
            country_name = country_names.get(s.get("country", ""), s.get("country", ""))
            where_val    = s.get("where", "")
            slug         = (what + "_" + where_val).replace(" ", "_")
            file_path    = os.path.join("jobs", slug, f"{slug}.xlsx")
            self.runs_tree.insert("", "end", iid=str(i),
                                  tags=("odd" if i % 2 else "even",),
                                  values=(what, s.get("where", ""), country_name, file_path))

    def _autorun_add(self):
        what    = self.run_what_var.get().strip()
        where   = self.run_where_var.get().strip()
        country = COUNTRIES.get(self.run_country_var.get(), "de")
        if not what:
            return
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        searches.append({"what": what, "where": where, "country": country})
        _save_config({"searches": searches})
        self.run_what_var.set("")
        self.run_where_var.set("")
        self._autorun_refresh()

    def _autorun_add_current(self):
        what    = self.what_var.get().strip()
        where   = self.where_var.get().strip()
        country = COUNTRIES.get(self.country_var.get(), "de")
        if not what:
            return
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        searches.append({"what": what, "where": where, "country": country})
        _save_config({"searches": searches})
        self._autorun_refresh()
        self.sched_status.config(
            text=f'Added: "{what}" in "{where}"', fg=SUCCESS)

    def _autorun_remove(self):
        sel = self.runs_tree.selection()
        if not sel:
            return
        idx      = int(sel[0])
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        if 0 <= idx < len(searches):
            searches.pop(idx)
            _save_config({"searches": searches})
            self._autorun_refresh()

    def _parse_log_history(self):
        """Return list of dicts for each individual search entry in the log (newest first)."""
        import re
        log_path = os.path.join(_app_dir(), "job_search.log")
        if not os.path.exists(log_path):
            return []

        entries = []
        current_dt = None
        pattern_search  = re.compile(r"Searching: '(.+)' in '(.+)' \(")
        pattern_fetched = re.compile(r">> (\d+) jobs fetched")
        pattern_result  = re.compile(r">> (\d+) new\s+\|\s+(\d+) duplicates skipped")
        pattern_time    = re.compile(r"\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]")

        pending = None
        with open(log_path, encoding="utf-8") as f:
            for line in f:
                m_time = pattern_time.match(line)
                if m_time:
                    current_dt = m_time.group(1)
                m_search = pattern_search.search(line)
                if m_search:
                    if pending:
                        entries.append(pending)
                    pending = {"datetime": current_dt,
                               "query": m_search.group(1),
                               "location": m_search.group(2),
                               "fetched": "", "new": "", "skipped": ""}
                    continue
                if pending:
                    m_f = pattern_fetched.search(line)
                    if m_f:
                        pending["fetched"] = m_f.group(1)
                        continue
                    m_r = pattern_result.search(line)
                    if m_r:
                        pending["new"]     = m_r.group(1)
                        pending["skipped"] = m_r.group(2)
                        entries.append(pending)
                        pending = None
        if pending:
            entries.append(pending)
        return list(reversed(entries))

    def _refresh_history(self):
        for row in self.hist_tree.get_children():
            self.hist_tree.delete(row)
        entries = self._parse_log_history()[:30]
        for i, e in enumerate(entries):
            self.hist_tree.insert("", "end", tags=("odd" if i % 2 else "even",),
                                  values=(e["datetime"], e["query"], e["location"],
                                          e["fetched"], e["new"], e["skipped"]))

    def _on_tab_change(self, idx=None):
        if idx is None:
            return
        if idx == 0:
            self._clear_btn.pack(side="right", padx=22)
        else:
            self._clear_btn.pack_forget()
        if idx == 1:
            self._refresh_saved_list()
        elif idx == 2:
            self._autorun_refresh()
            self._refresh_history()
        elif idx == 3:
            self._analytics_populate_files()

    # ── Analytics Panel ───────────────────────────────────────────────────────
    def _analytics_panel(self):
        # Scrollable container
        container = tk.Frame(self.tab_analytics, bg=BG)
        container.pack(fill="both", expand=True)
        canvas_scroll = tk.Canvas(container, bg=BG, highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical", command=canvas_scroll.yview)
        canvas_scroll.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas_scroll.pack(side="left", fill="both", expand=True)
        outer = tk.Frame(canvas_scroll, bg=BG)
        win_id = canvas_scroll.create_window((0, 0), window=outer, anchor="nw")
        outer.bind("<Configure>", lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all")))
        canvas_scroll.bind("<Configure>", lambda e: canvas_scroll.itemconfig(win_id, width=e.width))
        canvas_scroll.bind_all("<MouseWheel>", lambda e: canvas_scroll.yview_scroll(-1*(e.delta//120), "units"))

        hdr = tk.Frame(outer, bg=BG)
        hdr.pack(fill="x", padx=20, pady=(16, 8))
        tk.Label(hdr, text="Analytics", bg=BG, fg=ACCENT2,
                 font=("Segoe UI", 13, "bold")).pack(side="left")
        self._btn(hdr, "Export PDF", self._export_pdf_report,
                  color="#0f766e", w=12).pack(side="right", padx=(8, 0))
        self._btn(hdr, "Refresh", self._analytics_load,
                  color="#334155", w=10).pack(side="right")
        self._analytics_loading_lbl = tk.Label(hdr, text="", bg=BG, fg=SUBTEXT,
                                               font=("Segoe UI", 9, "italic"))
        self._analytics_loading_lbl.pack(side="left", expand=True)

        sel_row = tk.Frame(outer, bg=BG)
        sel_row.pack(fill="x", padx=20, pady=(0, 14))
        tk.Label(sel_row, text="File:", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 8))
        self.analytics_file_var = tk.StringVar(value="— select a file —")
        self.analytics_cb = ttk.Combobox(sel_row, textvariable=self.analytics_file_var,
                                         state="readonly", width=40)
        self.analytics_cb.pack(side="left")
        self.analytics_cb.bind("<<ComboboxSelected>>", lambda e: self._analytics_load())

        CH = 280  # chart height (2 per row = more space)

        # ── KPI cards ─────────────────────────────────────────────────────────
        kpi_row = tk.Frame(outer, bg=BG)
        kpi_row.pack(fill="x", padx=20, pady=(4, 20))

        def kpi_card(parent, val_attr, label_text):
            card = tk.Frame(parent, bg=PANEL, padx=16, pady=14)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8))
            val_lbl = tk.Label(card, text="—", bg=PANEL, fg=ACCENT2,
                               font=("Segoe UI", 18, "bold"))
            val_lbl.pack(anchor="center")
            tk.Label(card, text=label_text, bg=PANEL, fg=SUBTEXT,
                     font=("Segoe UI", 8)).pack(anchor="center")
            setattr(self, val_attr, val_lbl)

        kpi_card(kpi_row, "_kpi_jobs",    "Total Jobs")
        kpi_card(kpi_row, "_kpi_companies","Companies")
        kpi_card(kpi_row, "_kpi_salary",  "Avg Salary")
        kpi_card(kpi_row, "_kpi_range",   "Salary Range")
        kpi_card(kpi_row, "_kpi_cities",  "Cities")
        kpi_card(kpi_row, "_kpi_sources", "Sources")
        # last card no right margin
        kpi_card(kpi_row, "_kpi_weeks",   "Weeks Tracked")
        kpi_row.winfo_children()[-1].pack_configure(padx=0)

        def section_header(title):
            sec = tk.Frame(outer, bg=BG)
            sec.pack(fill="x", padx=20, pady=(16, 8))
            tk.Frame(sec, bg=BORDER, width=4, height=20).pack(side="left")
            tk.Label(sec, text=f"  {title}", bg=BG, fg=TEXT,
                     font=("Segoe UI", 10, "bold")).pack(side="left")
            tk.Frame(sec, bg=BORDER, height=1).pack(side="left", fill="x", expand=True, padx=(12, 0))

        def chart_pair(title_a, attr_a, stat_a, title_b, attr_b, stat_b):
            section_header(title_a + "  ·  " + title_b)
            row = tk.Frame(outer, bg=BG)
            row.pack(fill="x", padx=20, pady=(0, 4))
            for title, attr, stat in [(title_a, attr_a, stat_a), (title_b, attr_b, stat_b)]:
                last = (attr == attr_b)
                wrapper = tk.Frame(row, bg=PANEL)
                wrapper.pack(side="left", fill="both", expand=True,
                             padx=(0, 0 if last else 12))
                top = tk.Frame(wrapper, bg=PANEL)
                top.pack(fill="x", padx=14, pady=(12, 0))
                tk.Label(top, text=title, bg=PANEL, fg=ACCENT2,
                         font=("Segoe UI", 9, "bold")).pack(side="left")
                stat_lbl = tk.Label(top, text="", bg=PANEL, fg=SUBTEXT,
                                    font=("Segoe UI", 8))
                stat_lbl.pack(side="right")
                setattr(self, stat, stat_lbl)
                cf = tk.Frame(wrapper, bg=PANEL, height=CH)
                cf.pack(fill="both", expand=True, pady=(4, 12))
                cf.pack_propagate(False)
                setattr(self, attr, cf)

        chart_pair("Top Companies",      "chart_companies",    "_stat_companies",
                   "Jobs by Source",     "chart_source",       "_stat_source")
        chart_pair("Salary Distribution","chart_salary",       "_stat_salary",
                   "Avg Salary by City", "chart_salary_loc",   "_stat_salary_loc")
        chart_pair("Salary Trend",       "chart_salary_trend", "_stat_salary_trend",
                   "Top Keywords",       "chart_keywords",     "_stat_keywords")

        # Jobs Over Time — full width
        section_header("Jobs Added Over Time")
        trend_wrapper = tk.Frame(outer, bg=PANEL)
        trend_wrapper.pack(fill="x", padx=20, pady=(0, 24))
        top_trend = tk.Frame(trend_wrapper, bg=PANEL)
        top_trend.pack(fill="x", padx=14, pady=(12, 0))
        tk.Label(top_trend, text="Monthly Activity", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 9, "bold")).pack(side="left")
        self._stat_trend = tk.Label(top_trend, text="", bg=PANEL, fg=SUBTEXT,
                                    font=("Segoe UI", 8))
        self._stat_trend.pack(side="right")
        self.chart_trend = tk.Frame(trend_wrapper, bg=PANEL, height=180)
        self.chart_trend.pack(fill="x", pady=(4, 12))
        self.chart_trend.pack_propagate(False)

        self._analytics_populate_files()

    def _analytics_populate_files(self):
        jobs_dir = os.path.join(_app_dir(), "jobs")
        files = {}
        if os.path.isdir(jobs_dir):
            for root, _, fs in os.walk(jobs_dir):
                for f in fs:
                    if f.endswith(".xlsx"):
                        label = os.path.splitext(f)[0].replace("_", " ")
                        files[label] = os.path.join(root, f)
        self._analytics_files = files
        if not hasattr(self, "analytics_cb"):
            return
        self.analytics_cb["values"] = sorted(files.keys())
        if files:
            current = self.analytics_file_var.get()
            if current not in files:
                self.analytics_file_var.set(sorted(files.keys())[0])
            if self._active_tab == 3:
                self._analytics_load()

    def _analytics_load(self):
        label = self.analytics_file_var.get()
        path  = self._analytics_files.get(label)
        if not path or not os.path.exists(path):
            return

        # cancel any previous pending load
        if hasattr(self, "_analytics_thread") and self._analytics_thread.is_alive():
            self._analytics_cancel = True
        self._analytics_cancel = False

        if not hasattr(self, "_analytics_cache"):
            self._analytics_cache = {}
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            mtime = None

        # data cached and file unchanged → render in background, skip Excel loading
        cached = self._analytics_cache.get(path)
        if cached and cached["mtime"] == mtime:
            if hasattr(self, "_analytics_loading_lbl"):
                self._analytics_loading_lbl.config(text="Rendering…")
            if _PIL:
                d = cached["data"]
                def _render_cached():
                    imgs = self._render_charts_pil(d)
                    if not self._analytics_cancel:
                        self.after(0, lambda: self._display_pil_charts(imgs, d))
                t = threading.Thread(target=_render_cached, daemon=True)
                t.start()
            else:
                self.after(0, lambda: self._analytics_draw(**cached["data"]))
            return

        if hasattr(self, "_analytics_loading_lbl"):
            self._analytics_loading_lbl.config(text="Loading…")

        def _load():
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb["Job Results"]
                hdr = [str(c.value).strip() if c.value else ""
                       for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                wb.close()
            except Exception:
                return

            if self._analytics_cancel:
                return

            def col(name, fallback=-1):
                try: return hdr.index(name)
                except ValueError: return fallback

            i_company  = col("Company", 1)
            i_location = col("Location", 2)
            i_sal_min  = col("Salary Min (€)", 3)
            i_sal_max  = col("Salary Max (€)", 4)
            i_posted   = col("Posted", 5)
            i_desc     = col("Description", 7)
            i_source   = col("Source", -1)

            companies   = Counter()
            salaries    = []
            sal_by_loc  = defaultdict(list)
            sal_by_month= defaultdict(list)
            sources     = Counter()
            dates       = []
            all_words   = []
            STOPWORDS = {"the","and","or","for","with","that","this","from","are","you",
                         "our","your","will","have","has","can","its","all","more","than",
                         "also","their","they","which","who","what","when","where","how",
                         "about","new","use","using","used","work","we","be","been","but",
                         "not","they","as","in","on","at","to","of","a","an","is","it",
                         "und","die","der","das","mit","für","oder","bei","von","nicht",
                         "sind","wird","werden","haben","nach","auch","eine","einen","einem",
                         "wir","sie","sich","wie","durch","um","des","dem","den","im","an"}

            for r in rows:
                if self._analytics_cancel:
                    return
                if not r or len(r) < 3:
                    continue
                company  = str(r[i_company]  or "") if i_company  >= 0 else ""
                location = str(r[i_location] or "") if i_location >= 0 else ""
                sal_min  = r[i_sal_min] if i_sal_min >= 0 and i_sal_min < len(r) else None
                sal_max  = r[i_sal_max] if i_sal_max >= 0 and i_sal_max < len(r) else None
                posted   = str(r[i_posted] or "") if i_posted >= 0 and i_posted < len(r) else ""
                desc     = str(r[i_desc]   or "") if i_desc   >= 0 and i_desc   < len(r) else ""
                source   = str(r[i_source] or "") if i_source >= 0 and i_source < len(r) else ""

                if company: companies[company] += 1
                if source:  sources[source]    += 1

                sal = None
                if sal_min and sal_max:
                    try:
                        sal = (float(sal_min) + float(sal_max)) / 2
                        salaries.append(sal)
                        city = location.split(",")[0].strip()
                        if city: sal_by_loc[city].append(sal)
                    except Exception: pass

                if posted:
                    try:
                        dt = datetime.strptime(str(posted)[:10], "%Y-%m-%d")
                        dates.append(dt)
                        if sal:
                            sal_by_month[dt.strftime("%Y-%m")].append(sal)
                    except Exception: pass

                if desc:
                    words = [w.lower() for w in desc.split() if len(w) >= 4]
                    all_words.extend(w for w in words if w not in STOPWORDS and w.isalpha())

            keywords = Counter(all_words).most_common(15)

            if not self._analytics_cancel:
                data = dict(companies=companies, salaries=salaries,
                            sal_by_loc=sal_by_loc, sal_by_month=sal_by_month,
                            sources=sources, keywords=keywords, dates=dates)
                self._analytics_cache[path] = {"mtime": mtime, "data": data}
                if _PIL:
                    imgs = self._render_charts_pil(data)
                    if not self._analytics_cancel:
                        self.after(0, lambda d=data, i=imgs: self._display_pil_charts(i, d))
                else:
                    self.after(0, lambda: self._analytics_draw(**data))

        self._analytics_thread = threading.Thread(target=_load, daemon=True)
        self._analytics_thread.start()

    def _analytics_draw(self, companies, salaries, sal_by_loc, sal_by_month,
                        sources, keywords, dates, _labels_only=False):
        if not _labels_only:
            plt.close("all")
        self._analytics_data = dict(
            companies=companies, salaries=salaries, sal_by_loc=sal_by_loc,
            sal_by_month=sal_by_month, sources=sources, keywords=keywords, dates=dates,
            label=self.analytics_file_var.get(),
        )

        total_jobs = sum(companies.values())
        avg_sal    = sum(salaries) / len(salaries) if salaries else None
        months     = sorted(sal_by_month.keys())
        weeks      = len(set(d.strftime("%Y-%m") for d in dates)) if dates else 0
        top_src    = sources.most_common(1)

        # KPI cards
        self._kpi_jobs.config(text=f"{total_jobs:,}")
        self._kpi_companies.config(text=f"{len(companies):,}")
        self._kpi_salary.config(text=f"€{avg_sal:,.0f}" if avg_sal else "—")
        self._kpi_range.config(
            text=f"€{min(salaries):,.0f}–€{max(salaries):,.0f}" if salaries else "—")
        self._kpi_cities.config(text=f"{len(sal_by_loc):,}" if sal_by_loc else "—")
        self._kpi_sources.config(text=f"{len(sources):,}" if sources else "—")
        self._kpi_weeks.config(text=f"{weeks:,}" if weeks else "—")

        # chart stat labels
        self._stat_companies.config(text=f"{total_jobs} jobs")
        self._stat_salary.config(text=f"Ø €{avg_sal:,.0f}" if avg_sal else "no data")
        self._stat_salary_loc.config(text=f"{len(sal_by_loc)} cities" if sal_by_loc else "no data")
        self._stat_salary_trend.config(text=f"{months[0]} → {months[-1]}" if len(months) >= 2 else "")
        self._stat_source.config(text=f"top: {top_src[0][0]}" if top_src else "")
        self._stat_keywords.config(text=f"{len(keywords)} terms" if keywords else "")
        self._stat_trend.config(text=f"{len(dates)} points · {weeks} weeks" if dates else "")

        if _labels_only:
            return  # PIL path: images already displayed, only labels needed

        steps = [
            lambda: self._draw_companies(companies),
            lambda: self._draw_salary(salaries),
            lambda: self._draw_salary_loc(sal_by_loc),
            lambda: self._draw_salary_trend(sal_by_month),
            lambda: self._draw_source(sources),
            lambda: self._draw_keywords(keywords),
            lambda: self._draw_trend(dates),
        ]
        total = len(steps)

        def draw_next(i=0):
            if self._analytics_cancel:
                return
            if i >= total:
                if hasattr(self, "_analytics_loading_lbl"):
                    self._analytics_loading_lbl.config(text="")
                return
            if hasattr(self, "_analytics_loading_lbl"):
                self._analytics_loading_lbl.config(text=f"Drawing… {i+1}/{total}")
            steps[i]()
            self.after(5, lambda: draw_next(i + 1))

        draw_next()

    def _render_charts_pil(self, data):
        """Render all charts to PIL Images using Agg backend (background-thread safe)."""
        companies    = data["companies"]
        salaries     = data["salaries"]
        sal_by_loc   = data["sal_by_loc"]
        sal_by_month = data["sal_by_month"]
        sources      = data["sources"]
        keywords     = data["keywords"]
        dates        = data["dates"]

        BG = "#1e293b"; AX = "#0f172a"; SUB = "#94a3b8"; FG = "#f1f5f9"; GRD = "#334155"

        def mf(w=5.5, h=2.8):
            fig = Figure(figsize=(w, h), facecolor=BG)
            FigureCanvasAgg(fig)
            ax = fig.add_subplot(111)
            ax.set_facecolor(AX)
            ax.tick_params(colors=SUB, labelsize=8)
            for sp in ax.spines.values(): sp.set_color(GRD)
            ax.title.set_color(FG); ax.xaxis.label.set_color(SUB); ax.yaxis.label.set_color(SUB)
            return fig, ax

        def topil(fig):
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=100, bbox_inches="tight",
                        facecolor=fig.get_facecolor())
            buf.seek(0)
            img = Image.open(buf).copy()
            buf.close()
            return img

        imgs = {}

        # Companies
        fig, ax = mf()
        top = companies.most_common(10)
        if top:
            names = [c[0][:22] for c in reversed(top)]
            cnts  = [c[1] for c in reversed(top)]
            bars = ax.barh(names, cnts, color="#0ea5e9", height=0.6)
            ax.bar_label(bars, fmt="%d", color=FG, fontsize=7, padding=3)
            ax.set_title("Top Companies", fontsize=10, pad=8)
            ax.set_xlabel("Jobs", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["companies"] = topil(fig)

        # Salary distribution
        fig, ax = mf()
        if salaries:
            avg = sum(salaries) / len(salaries)
            ax.hist(salaries, bins=12, color="#38bdf8", edgecolor=AX, linewidth=0.5)
            ax.axvline(avg, color="#34d399", linewidth=1.5, linestyle="--", label=f"Avg: {avg:,.0f} €")
            ax.legend(fontsize=7, facecolor=BG, edgecolor=GRD, labelcolor=FG)
            ax.set_title("Salary Distribution", fontsize=10, pad=8)
            ax.set_xlabel("Salary (€)", fontsize=8); ax.set_ylabel("Jobs", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["salary"] = topil(fig)

        # Salary by location
        fig, ax = mf()
        filtered = {k: v for k, v in sal_by_loc.items() if v}
        if filtered:
            avg_loc = {k: sum(v)/len(v) for k, v in filtered.items()}
            top_loc = sorted(avg_loc.items(), key=lambda x: x[1], reverse=True)[:10]
            names = [t[0][:18] for t in reversed(top_loc)]
            avgs  = [t[1] for t in reversed(top_loc)]
            bars = ax.barh(names, avgs, color="#a78bfa", height=0.6)
            ax.bar_label(bars, labels=[f"{v:,.0f}" for v in avgs], color=FG, fontsize=7, padding=3)
            ax.set_title("Avg Salary by Location", fontsize=10, pad=8)
            ax.set_xlabel("Avg Salary (€)", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["salary_loc"] = topil(fig)

        # Salary trend
        fig, ax = mf()
        if sal_by_month:
            months = sorted(sal_by_month.keys())
            avgs   = [sum(sal_by_month[m])/len(sal_by_month[m]) for m in months]
            mdts   = [datetime.strptime(m, "%Y-%m") for m in months]
            ax.plot(mdts, avgs, color="#34d399", linewidth=2, marker="o",
                    markersize=5, markerfacecolor="#34d399")
            ax.fill_between(mdts, avgs, alpha=0.2, color="#34d399")
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
            for lbl in ax.get_xticklabels(): lbl.set_rotation(30); lbl.set_ha("right")
            ax.set_title("Salary Trend", fontsize=10, pad=8)
            ax.set_ylabel("Avg Salary (€)", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["salary_trend"] = topil(fig)

        # Source
        fig, ax = mf()
        if sources:
            top_s  = sources.most_common(8)
            names  = [s[0][:16] for s in reversed(top_s)]
            counts = [s[1] for s in reversed(top_s)]
            bars = ax.barh(names, counts, color="#fb923c", height=0.6)
            ax.bar_label(bars, fmt="%d", color=FG, fontsize=7, padding=3)
            ax.set_title("Jobs by Source", fontsize=10, pad=8)
            ax.set_xlabel("Jobs", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["source"] = topil(fig)

        # Keywords
        fig, ax = mf(w=5.5, h=3.0)
        if keywords:
            words  = [k[0][:20] for k in reversed(keywords)]
            counts = [k[1] for k in reversed(keywords)]
            bars = ax.barh(words, counts, color="#818cf8", height=0.6)
            ax.bar_label(bars, fmt="%d", color=FG, fontsize=7, padding=3)
            ax.set_title("Top Keywords", fontsize=10, pad=8)
            ax.set_xlabel("Occurrences", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["keywords"] = topil(fig)

        # Trend over time
        fig, ax = mf(w=11, h=1.8)
        if dates:
            weekly = defaultdict(int)
            for d in dates: weekly[d.strftime("%Y-%m")] += 1
            sw = sorted(weekly.keys())
            wdates = [datetime.strptime(w, "%Y-%m") for w in sw]
            cnts   = [weekly[w] for w in sw]
            ax.fill_between(wdates, cnts, color="#0ea5e9", alpha=0.3)
            ax.plot(wdates, cnts, color="#38bdf8", linewidth=1.5,
                    marker="o", markersize=4, markerfacecolor="#0ea5e9")
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            for lbl in ax.get_xticklabels(): lbl.set_rotation(30); lbl.set_ha("right")
            ax.set_ylabel("New Jobs", fontsize=8)
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB, transform=ax.transAxes)
        fig.tight_layout(); imgs["trend"] = topil(fig)

        return imgs

    def _display_pil_charts(self, imgs, data):
        """Display PIL images in chart frames + update labels. Runs in main thread."""
        if not hasattr(self, "_pil_refs"):
            self._pil_refs = {}

        def show(frame, key):
            img = imgs.get(key)
            if img is None:
                return
            for w in frame.winfo_children():
                w.destroy()
            # resize to fit frame
            fw = frame.winfo_width()  or img.width
            fh = frame.winfo_height() or img.height
            img_r = img.copy()
            img_r.thumbnail((fw, fh), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img_r)
            self._pil_refs[key] = photo
            lbl = tk.Label(frame, image=photo, bg="#1e293b", anchor="center")
            lbl.pack(fill="both", expand=True)

        show(self.chart_companies,    "companies")
        show(self.chart_salary,       "salary")
        show(self.chart_salary_loc,   "salary_loc")
        show(self.chart_salary_trend, "salary_trend")
        show(self.chart_source,       "source")
        show(self.chart_keywords,     "keywords")
        show(self.chart_trend,        "trend")

        if hasattr(self, "_analytics_loading_lbl"):
            self._analytics_loading_lbl.config(text="")

        self._analytics_draw(**data, _labels_only=True)  # update KPI + stat labels only

    def _export_pdf_report(self):
        from matplotlib.backends.backend_pdf import PdfPages
        if not hasattr(self, "_analytics_data") or not self._analytics_data:
            messagebox.showinfo("Export PDF", "Please load analytics data first.")
            return
        d = self._analytics_data
        companies   = d["companies"]
        salaries    = d["salaries"]
        sal_by_loc  = d["sal_by_loc"]
        sal_by_month= d["sal_by_month"]
        sources     = d["sources"]
        keywords    = d["keywords"]
        dates       = d["dates"]
        label       = d["label"]

        default_name = label.replace(" ", "_") + "_report.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=default_name,
            title="Save PDF Report",
        )
        if not path:
            return

        BG_P   = "#1e293b"
        FG_P   = "#f1f5f9"
        SUB_P  = "#94a3b8"
        ACC_P  = "#38bdf8"
        AX_P   = "#0f172a"
        GRID_P = "#334155"

        def _style(fig, ax):
            fig.patch.set_facecolor(BG_P)
            ax.set_facecolor(AX_P)
            ax.tick_params(colors=SUB_P, labelsize=8)
            ax.spines[:].set_color(GRID_P)
            ax.title.set_color(FG_P)
            ax.xaxis.label.set_color(SUB_P)
            ax.yaxis.label.set_color(SUB_P)

        def _page_footer(fig, label, page):
            fig.text(0.5, 0.01, f"{label}  ·  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  Page {page}",
                     ha="center", va="bottom", fontsize=7, color=SUB_P)

        def _insight_box(fig, y, text):
            fig.text(0.07, y, text, fontsize=8.5, color=FG_P,
                     va="top", wrap=True,
                     bbox=dict(boxstyle="round,pad=0.5", facecolor="#0f172a",
                               edgecolor=GRID_P, linewidth=0.8),
                     transform=fig.transFigure,
                     horizontalalignment="left",
                     multialignment="left",
                     figure=fig)

        total_jobs = sum(companies.values())
        avg_sal    = sum(salaries) / len(salaries) if salaries else 0
        top5_co    = companies.most_common(5)
        top_src    = sources.most_common(3)
        top_kw     = [k for k, _ in keywords[:5]]
        months     = sorted(sal_by_month.keys())

        with PdfPages(path) as pdf:
            # ── PAGE 1: Cover ─────────────────────────────────────────────────
            fig = plt.figure(figsize=(11, 8.5))
            fig.patch.set_facecolor(BG_P)

            fig.text(0.5, 0.82, "Job Market Analytics Report",
                     ha="center", fontsize=26, fontweight="bold", color=ACC_P)
            fig.text(0.5, 0.76, label,
                     ha="center", fontsize=14, color=FG_P)
            fig.text(0.5, 0.71, datetime.now().strftime("%B %d, %Y"),
                     ha="center", fontsize=10, color=SUB_P)

            # Summary stats grid
            stats = [
                ("Total Jobs",     f"{total_jobs:,}"),
                ("Companies",      f"{len(companies):,}"),
                ("With Salary",    f"{len(salaries):,}"),
                ("Avg Salary",     f"€{avg_sal:,.0f}" if salaries else "—"),
                ("Cities",         f"{len(sal_by_loc):,}"),
                ("Sources",        f"{len(sources):,}"),
                ("Months Tracked", f"{len(set(d.strftime('%Y-%m') for d in dates)):,}" if dates else "—"),
                ("Top Keywords",   ", ".join(top_kw[:3]) if top_kw else "—"),
            ]
            cols = 4
            for i, (lbl, val) in enumerate(stats):
                x = 0.08 + (i % cols) * 0.23
                y = 0.56 - (i // cols) * 0.13
                ax_s = fig.add_axes([x, y, 0.20, 0.10])
                ax_s.set_facecolor("#0f172a")
                ax_s.set_xticks([]); ax_s.set_yticks([])
                for sp in ax_s.spines.values(): sp.set_color(GRID_P)
                ax_s.text(0.5, 0.65, val, ha="center", va="center",
                          fontsize=14, fontweight="bold", color=ACC_P,
                          transform=ax_s.transAxes)
                ax_s.text(0.5, 0.20, lbl, ha="center", va="center",
                          fontsize=8, color=SUB_P, transform=ax_s.transAxes)

            # Key findings text
            findings = []
            if top5_co:
                top_name, top_cnt = top5_co[0]
                pct = top_cnt / total_jobs * 100 if total_jobs else 0
                findings.append(f"• Most active recruiter: {top_name} with {top_cnt} postings ({pct:.1f}% of all jobs).")
            if salaries:
                findings.append(f"• Average salary: €{avg_sal:,.0f}  |  Range: €{min(salaries):,.0f} – €{max(salaries):,.0f}.")
            if months and len(months) >= 2:
                first_avg = sum(sal_by_month[months[0]]) / len(sal_by_month[months[0]])
                last_avg  = sum(sal_by_month[months[-1]]) / len(sal_by_month[months[-1]])
                direction = "rose" if last_avg > first_avg else "fell"
                findings.append(f"• Salary {direction} from €{first_avg:,.0f} ({months[0]}) to €{last_avg:,.0f} ({months[-1]}).")
            if top_src:
                findings.append(f"• Top source: {top_src[0][0]} ({top_src[0][1]} jobs).")
            if top_kw:
                findings.append(f"• Most demanded skills: {', '.join(top_kw)}.")

            fig.text(0.08, 0.30, "Key Findings", fontsize=11, fontweight="bold",
                     color=ACC_P, va="top")
            fig.text(0.08, 0.25, "\n".join(findings), fontsize=9, color=FG_P,
                     va="top", linespacing=1.8)

            _page_footer(fig, label, 1)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

            def _two_chart_page(page_num, draw_left, draw_right, ins_left, ins_right):
                """Page with 2 charts + dedicated text box per chart — no overlaps."""
                from matplotlib.gridspec import GridSpec
                fig = plt.figure(figsize=(11, 8.5))
                fig.patch.set_facecolor(BG_P)
                gs  = GridSpec(2, 2, figure=fig,
                               height_ratios=[3.2, 1],
                               left=0.07, right=0.97,
                               top=0.92, bottom=0.10,
                               hspace=0.45, wspace=0.35)
                ax_l  = fig.add_subplot(gs[0, 0])
                ax_r  = fig.add_subplot(gs[0, 1])
                txt_l = fig.add_subplot(gs[1, 0])
                txt_r = fig.add_subplot(gs[1, 1])
                _style(fig, ax_l); _style(fig, ax_r)
                draw_left(ax_l); draw_right(ax_r)
                import textwrap
                for txt_ax, ins in ((txt_l, ins_left), (txt_r, ins_right)):
                    txt_ax.set_facecolor("#0f172a")
                    txt_ax.set_xticks([]); txt_ax.set_yticks([])
                    for sp in txt_ax.spines.values(): sp.set_color(GRID_P)
                    wrapped = textwrap.fill(ins, width=62)
                    txt_ax.text(0.03, 0.92, wrapped, va="top", ha="left",
                                fontsize=8, color=FG_P,
                                transform=txt_ax.transAxes,
                                linespacing=1.5)
                _page_footer(fig, label, page_num)
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

            # ── PAGE 2: Top Companies + Salary Distribution ────────────────────
            top5_str = ", ".join(f"{n} ({c})" for n, c in top5_co) if top5_co else "—"
            top5_pct = sum(c for _, c in top5_co) / total_jobs * 100 if total_jobs else 0
            ins_co   = (f"Top Companies: {len(companies)} unique employers found across "
                        f"{total_jobs} postings. The top 5 — {top5_str} — account for "
                        f"{top5_pct:.0f}% of all listings.")
            if salaries:
                median  = sorted(salaries)[len(salaries)//2]
                ins_sal = (f"Salary Distribution: {len(salaries)} jobs with salary data. "
                           f"Average: €{avg_sal:,.0f}, median: €{median:,.0f}. "
                           f"Range: €{min(salaries):,.0f} – €{max(salaries):,.0f}.")
            else:
                ins_sal = "Salary Distribution: No salary data available."

            def _draw_companies(ax):
                if companies:
                    top_n = companies.most_common(10)
                    names, cnts = zip(*reversed(top_n))
                    bars = ax.barh(list(names), list(cnts), color="#818cf8", height=0.6)
                    ax.bar_label(bars, fmt="%d", color=FG_P, fontsize=7, padding=3)
                    ax.set_title("Top Companies", fontsize=11, pad=10)
                    ax.set_xlabel("Job Postings", fontsize=8)
                else:
                    ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB_P)

            def _draw_salary_dist(ax):
                if salaries:
                    ax.hist(salaries, bins=20, color="#34d399", edgecolor=AX_P, linewidth=0.4)
                    ax.axvline(avg_sal, color="#f87171", linewidth=1.5, linestyle="--",
                               label=f"Avg €{avg_sal:,.0f}")
                    ax.legend(fontsize=8, labelcolor=FG_P, facecolor=AX_P, edgecolor=GRID_P)
                    ax.set_title("Salary Distribution", fontsize=11, pad=10)
                    ax.set_xlabel("Annual Salary (€)", fontsize=8)
                    ax.set_ylabel("Number of Jobs", fontsize=8)
                else:
                    ax.text(0.5, 0.5, "No salary data", ha="center", va="center", color=SUB_P)

            _two_chart_page(2, _draw_companies, _draw_salary_dist, ins_co, ins_sal)

            # ── PAGE 3: Avg Salary by City + Salary Trend ─────────────────────
            city_avg = {}
            if sal_by_loc:
                city_avg   = {c: sum(v)/len(v) for c, v in sal_by_loc.items() if v}
                top_cities = sorted(city_avg, key=city_avg.get, reverse=True)[:10]
                best_city  = max(city_avg, key=city_avg.get)
                worst_city = min(city_avg, key=city_avg.get)
                ins_loc = (f"Avg Salary by City: {len(sal_by_loc)} cities analysed. "
                           f"Highest: {best_city} (€{city_avg[best_city]:,.0f}). "
                           f"Lowest: {worst_city} (€{city_avg[worst_city]:,.0f}).")
            else:
                ins_loc = "Avg Salary by City: No location/salary data available."
            if months and len(months) >= 2:
                first_avg = sum(sal_by_month[months[0]])/len(sal_by_month[months[0]])
                last_avg  = sum(sal_by_month[months[-1]])/len(sal_by_month[months[-1]])
                chg = (last_avg - first_avg) / first_avg * 100 if first_avg else 0
                ins_trend = (f"Salary Trend: {months[0]} to {months[-1]} ({len(months)} months). "
                             f"Salary {'increased' if chg >= 0 else 'decreased'} by "
                             f"{abs(chg):.1f}% (€{first_avg:,.0f} → €{last_avg:,.0f}).")
            else:
                ins_trend = "Salary Trend: Not enough monthly data."

            def _draw_city(ax):
                if sal_by_loc:
                    avgs = [city_avg[c] for c in reversed(top_cities)]
                    bars = ax.barh(list(reversed(top_cities)), avgs, color="#f59e0b", height=0.6)
                    ax.bar_label(bars, labels=[f"€{v:,.0f}" for v in avgs],
                                 color=FG_P, fontsize=7, padding=3)
                    ax.set_title("Avg Salary by City", fontsize=11, pad=10)
                    ax.set_xlabel("Average Salary (€)", fontsize=8)
                else:
                    ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB_P)

            def _draw_trend(ax):
                if sal_by_month and len(sal_by_month) >= 2:
                    ms     = sorted(sal_by_month.keys())
                    avgs_m = [sum(sal_by_month[m])/len(sal_by_month[m]) for m in ms]
                    ax.plot(ms, avgs_m, color="#38bdf8", linewidth=2,
                            marker="o", markersize=5, markerfacecolor="#0ea5e9")
                    ax.fill_between(ms, avgs_m, alpha=0.15, color="#38bdf8")
                    ax.set_title("Salary Trend (monthly avg)", fontsize=11, pad=10)
                    ax.set_ylabel("Avg Salary (€)", fontsize=8)
                    plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right", fontsize=7)
                else:
                    ax.text(0.5, 0.5, "Not enough data", ha="center", va="center", color=SUB_P)

            _two_chart_page(3, _draw_city, _draw_trend, ins_loc, ins_trend)

            # ── PAGE 4: Jobs by Source + Top Keywords ──────────────────────────
            src_items = sources.most_common() if sources else []
            if src_items:
                src_names, src_cnts = zip(*src_items)
                src_pct  = {n: c/total_jobs*100 for n, c in src_items} if total_jobs else {}
                ins_src  = (f"Jobs by Source: {len(sources)} sources. "
                            f"{src_names[0]} leads with {src_cnts[0]} jobs "
                            f"({src_pct.get(src_names[0],0):.0f}%).")
                if len(src_names) > 1:
                    ins_src += f" Followed by {src_names[1]} ({src_cnts[1]} jobs)."
            else:
                src_names = src_cnts = []
                ins_src = "Jobs by Source: No source data recorded."
            if keywords:
                kw_top = [k for k, _ in keywords[:8]]
                ins_kw = (f"Top Keywords: Most frequent terms: {', '.join(kw_top[:5])}. "
                          f"These reflect the core skills employers are seeking.")
            else:
                ins_kw = "Top Keywords: No description data available."

            def _draw_sources(ax):
                if src_items:
                    colors_src = ["#38bdf8","#818cf8","#34d399","#f59e0b",
                                  "#f87171","#a78bfa","#fb923c","#4ade80"][:len(src_names)]
                    _, _, autotexts = ax.pie(
                        src_cnts, labels=src_names, colors=colors_src,
                        autopct="%1.0f%%", startangle=90,
                        textprops={"color": FG_P, "fontsize": 8})
                    for at in autotexts: at.set_fontsize(8)
                    ax.set_title("Jobs by Source", fontsize=11, pad=10)
                else:
                    ax.text(0.5, 0.5, "No source data", ha="center", va="center", color=SUB_P)

            def _draw_keywords(ax):
                if keywords:
                    kwords, kcounts = zip(*reversed(keywords[:15]))
                    bars = ax.barh(list(kwords), list(kcounts), color="#818cf8", height=0.6)
                    ax.bar_label(bars, fmt="%d", color=FG_P, fontsize=7, padding=3)
                    ax.set_title("Top Keywords in Job Descriptions", fontsize=11, pad=10)
                    ax.set_xlabel("Occurrences", fontsize=8)
                else:
                    ax.text(0.5, 0.5, "No data", ha="center", va="center", color=SUB_P)

            _two_chart_page(4, _draw_sources, _draw_keywords, ins_src, ins_kw)

            # ── PAGE 5: Jobs Over Time ─────────────────────────────────────────
            from matplotlib.gridspec import GridSpec as GS5
            fig = plt.figure(figsize=(11, 8.5))
            fig.patch.set_facecolor(BG_P)
            gs5 = GS5(2, 1, figure=fig, height_ratios=[3.2, 1],
                      left=0.07, right=0.97, top=0.92, bottom=0.10, hspace=0.45)
            ax      = fig.add_subplot(gs5[0, 0]); _style(fig, ax)
            txt_ax  = fig.add_subplot(gs5[1, 0])

            if dates:
                weekly = defaultdict(int)
                for dt in dates: weekly[dt.strftime("%Y-%m")] += 1
                sorted_weeks = sorted(weekly.keys())
                week_dates   = [datetime.strptime(w, "%Y-%m") for w in sorted_weeks]
                counts       = [weekly[w] for w in sorted_weeks]
                ax.fill_between(week_dates, counts, color="#0ea5e9", alpha=0.3)
                ax.plot(week_dates, counts, color="#38bdf8", linewidth=2,
                        marker="o", markersize=4, markerfacecolor="#0ea5e9")
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
                ax.xaxis.set_major_locator(mdates.MonthLocator())
                plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right", fontsize=7)
                ax.set_ylabel("New Jobs per Month", fontsize=8)
                peak_month = sorted_weeks[counts.index(max(counts))]
                ins_time = (f"Jobs Over Time: {len(dates)} postings tracked across "
                            f"{len(sorted_weeks)} months. Peak: {peak_month} "
                            f"({max(counts)} jobs). Avg {sum(counts)/len(counts):.0f} jobs/month.")
            else:
                ax.text(0.5, 0.5, "No date data", ha="center", va="center", color=SUB_P)
                ins_time = "Jobs Over Time: No date data available."

            ax.set_title("Jobs Added Over Time", fontsize=12, pad=10)
            txt_ax.set_facecolor("#0f172a")
            txt_ax.set_xticks([]); txt_ax.set_yticks([])
            for sp in txt_ax.spines.values(): sp.set_color(GRID_P)
            import textwrap
            wrapped_time = textwrap.fill(ins_time, width=90)
            txt_ax.text(0.03, 0.92, wrapped_time, va="top", ha="left",
                        fontsize=8, color=FG_P,
                        transform=txt_ax.transAxes, linespacing=1.5)
            _page_footer(fig, label, 5)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

        messagebox.showinfo("Export PDF", f"Report saved:\n{path}")

    def _make_fig(self, frame, figsize=(4, 3)):
        for w in frame.winfo_children():
            w.destroy()
        fig, ax = plt.subplots(figsize=figsize)
        fig.patch.set_facecolor("#1e293b")
        ax.set_facecolor("#0f172a")
        ax.tick_params(colors="#94a3b8", labelsize=8)
        ax.spines[:].set_color("#334155")
        ax.title.set_color("#f1f5f9")
        ax.xaxis.label.set_color("#94a3b8")
        ax.yaxis.label.set_color("#94a3b8")
        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        return fig, ax, canvas

    def _no_data(self, ax, canvas, title):
        ax.set_title(title, fontsize=10, pad=8)
        ax.text(0.5, 0.5, "No data", ha="center", va="center",
                color="#94a3b8", transform=ax.transAxes)
        canvas.draw()

    def _draw_companies(self, companies):
        fig, ax, canvas = self._make_fig(self.chart_companies, figsize=(4, 3.5))
        top = companies.most_common(10)
        if not top: return self._no_data(ax, canvas, "Top Companies")
        names  = [c[0][:20] for c in reversed(top)]
        counts = [c[1]       for c in reversed(top)]
        bars = ax.barh(names, counts, color="#0ea5e9", height=0.6)
        ax.bar_label(bars, fmt="%d", color="#f1f5f9", fontsize=7, padding=3)
        ax.set_title("Top Companies", fontsize=10, pad=8)
        ax.set_xlabel("Jobs", fontsize=8)
        plt.tight_layout(); canvas.draw()

    def _draw_salary(self, salaries):
        fig, ax, canvas = self._make_fig(self.chart_salary, figsize=(4, 3.5))
        if not salaries: return self._no_data(ax, canvas, "Salary Distribution")
        ax.hist(salaries, bins=12, color="#38bdf8", edgecolor="#0f172a", linewidth=0.5)
        avg = sum(salaries) / len(salaries)
        ax.axvline(avg, color="#34d399", linewidth=1.5, linestyle="--", label=f"Avg: {avg:,.0f} €")
        ax.legend(fontsize=7, facecolor="#1e293b", edgecolor="#334155", labelcolor="#f1f5f9")
        ax.set_title("Salary Distribution", fontsize=10, pad=8)
        ax.set_xlabel("Salary (€)", fontsize=8)
        ax.set_ylabel("Jobs", fontsize=8)
        plt.tight_layout(); canvas.draw()

    def _draw_salary_loc(self, sal_by_loc):
        fig, ax, canvas = self._make_fig(self.chart_salary_loc, figsize=(4, 3.5))
        filtered = {k: v for k, v in sal_by_loc.items() if len(v) >= 1}
        if not filtered: return self._no_data(ax, canvas, "Salary by Location")
        avg_by_loc = {k: sum(v)/len(v) for k, v in filtered.items()}
        top = sorted(avg_by_loc.items(), key=lambda x: x[1], reverse=True)[:10]
        names = [t[0][:18] for t in reversed(top)]
        avgs  = [t[1]       for t in reversed(top)]
        bars = ax.barh(names, avgs, color="#a78bfa", height=0.6)
        ax.bar_label(bars, labels=[f"{v:,.0f}" for v in avgs],
                     color="#f1f5f9", fontsize=7, padding=3)
        ax.set_title("Avg Salary by Location", fontsize=10, pad=8)
        ax.set_xlabel("Avg Salary (€)", fontsize=8)
        plt.tight_layout(); canvas.draw()

    def _draw_salary_trend(self, sal_by_month):
        fig, ax, canvas = self._make_fig(self.chart_salary_trend, figsize=(4, 3.5))
        if not sal_by_month: return self._no_data(ax, canvas, "Salary Trend")
        months = sorted(sal_by_month.keys())
        avgs   = [sum(sal_by_month[m])/len(sal_by_month[m]) for m in months]
        month_dates = [datetime.strptime(m, "%Y-%m") for m in months]
        ax.plot(month_dates, avgs, color="#34d399", linewidth=2, marker="o",
                markersize=5, markerfacecolor="#34d399")
        ax.fill_between(month_dates, avgs, alpha=0.2, color="#34d399")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right", fontsize=7)
        ax.set_title("Salary Trend", fontsize=10, pad=8)
        ax.set_ylabel("Avg Salary (€)", fontsize=8)
        plt.tight_layout(); canvas.draw()

    def _draw_source(self, sources):
        fig, ax, canvas = self._make_fig(self.chart_source, figsize=(4, 3.5))
        if not sources: return self._no_data(ax, canvas, "Jobs by Source")
        top = sources.most_common(8)
        names  = [s[0][:16] for s in reversed(top)]
        counts = [s[1]       for s in reversed(top)]
        bars = ax.barh(names, counts, color="#fb923c", height=0.6)
        ax.bar_label(bars, fmt="%d", color="#f1f5f9", fontsize=7, padding=3)
        ax.set_title("Jobs by Source", fontsize=10, pad=8)
        ax.set_xlabel("Jobs", fontsize=8)
        plt.tight_layout(); canvas.draw()

    def _draw_keywords(self, keywords):
        fig, ax, canvas = self._make_fig(self.chart_keywords, figsize=(5.5, 3.5))
        if not keywords: return self._no_data(ax, canvas, "Top Keywords")
        words  = [k[0][:20] for k in reversed(keywords)]
        counts = [k[1]       for k in reversed(keywords)]
        bars = ax.barh(words, counts, color="#818cf8", height=0.6)
        ax.bar_label(bars, fmt="%d", color="#f1f5f9", fontsize=7, padding=3)
        ax.set_title("Top Keywords in Job Descriptions", fontsize=10, pad=8)
        ax.set_xlabel("Occurrences", fontsize=8)
        plt.tight_layout(); canvas.draw()

    def _draw_trend(self, dates):
        fig, ax, canvas = self._make_fig(self.chart_trend, figsize=(10, 2.2))
        if not dates: return self._no_data(ax, canvas, "Jobs Added Over Time")
        weekly = defaultdict(int)
        for d in dates:
            weekly[d.strftime("%Y-%m")] += 1
        sorted_weeks = sorted(weekly.keys())
        week_dates = [datetime.strptime(w, "%Y-%m") for w in sorted_weeks]
        counts     = [weekly[w] for w in sorted_weeks]
        ax.fill_between(week_dates, counts, color="#0ea5e9", alpha=0.3)
        ax.plot(week_dates, counts, color="#38bdf8", linewidth=1.5, marker="o",
                markersize=4, markerfacecolor="#0ea5e9")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right", fontsize=7)
        ax.set_title("Jobs Added Over Time", fontsize=10, pad=8)
        ax.set_ylabel("New Jobs", fontsize=8)
        plt.tight_layout(); canvas.draw()

    # ── Credentials Panel ─────────────────────────────────────────────────────
    def _credentials_panel(self):
        cfg = _load_config()

        def make_step(parent, num, text, padx=20):
            row = tk.Frame(parent, bg=PANEL)
            row.pack(anchor="w", padx=padx, pady=2)
            tk.Label(row, text=num, bg=ACCENT, fg="white",
                     font=("Segoe UI", 8, "bold"), width=2,
                     relief="flat").pack(side="left", padx=(0, 10))
            tk.Label(row, text=text, bg=PANEL, fg=TEXT,
                     font=("Segoe UI", 10), wraplength=340,
                     justify="left").pack(side="left", anchor="w")

        # ── Custom sub-tab bar (same style as main nav) ───────────────────────
        _cred_names = ["Adzuna", "Reed", "Findwork", "Jooble", "HeadHunter"]
        _cred_active = [0]
        _cred_btns   = []
        _cred_inds   = []
        _cred_frames = []

        # ── BYOK privacy notice ───────────────────────────────────────────────
        byok_bar = tk.Frame(self.tab_credentials, bg="#0f2744")
        byok_bar.pack(fill="x")
        tk.Label(byok_bar,
                 text="🔒  Your API keys are saved locally on this device only — we never transmit or store them externally.",
                 bg="#0f2744", fg="#93c5fd",
                 font=("Segoe UI", 9), pady=6).pack(anchor="w", padx=20)

        sub_bar_wrap = tk.Frame(self.tab_credentials, bg=BG)
        sub_bar_wrap.pack(fill="x")
        tk.Frame(sub_bar_wrap, bg=ACCENT, height=1).pack(fill="x", side="top")
        sub_btn_row = tk.Frame(sub_bar_wrap, bg=BG)
        sub_btn_row.pack(fill="x", side="top")

        content_area = tk.Frame(self.tab_credentials, bg=BG)
        content_area.pack(fill="both", expand=True)

        def _cred_switch(idx):
            _cred_active[0] = idx
            for i, (b, ind, f) in enumerate(zip(_cred_btns, _cred_inds, _cred_frames)):
                if i == idx:
                    b.config(fg=ACCENT, font=("Segoe UI", 10, "bold"))
                    ind.config(bg=ACCENT)
                    f.pack(fill="both", expand=True)
                else:
                    b.config(fg=SUBTEXT, font=("Segoe UI", 10))
                    ind.config(bg=BG)
                    f.pack_forget()

        for i, name in enumerate(_cred_names):
            col = tk.Frame(sub_btn_row, bg=BG)
            col.pack(side="left")
            btn = tk.Label(col, text=name, bg=BG,
                           fg=ACCENT if i == 0 else SUBTEXT,
                           font=("Segoe UI", 10, "bold" if i == 0 else "normal"),
                           padx=16, pady=8, cursor="hand2")
            btn.pack()
            ind = tk.Frame(col, height=2, bg=ACCENT if i == 0 else BG)
            ind.pack(fill="x")
            _cred_btns.append(btn)
            _cred_inds.append(ind)
            btn.bind("<Button-1>", lambda e, idx=i: _cred_switch(idx))
            btn.bind("<Enter>",  lambda e, b=btn, idx=i: (
                b.config(fg=ACCENT2) if idx != _cred_active[0] else None))
            btn.bind("<Leave>",  lambda e, *_: _cred_switch(_cred_active[0]))

        tk.Frame(sub_bar_wrap, bg=ACCENT, height=1).pack(fill="x", side="bottom")

        def _make_cred_tab():
            f = tk.Frame(content_area, bg=BG)
            _cred_frames.append(f)
            return f

        tab_az      = _make_cred_tab()
        tab_reed    = _make_cred_tab()
        tab_cj      = _make_cred_tab()
        tab_jooble  = _make_cred_tab()
        tab_hh      = _make_cred_tab()

        # ── Adzuna sub-tab ────────────────────────────────────────────────────
        az = tk.Frame(tab_az, bg=PANEL)
        az.pack(fill="both", expand=True, padx=30, pady=20)

        az_left = tk.Frame(az, bg=PANEL)
        az_left.pack(side="left", fill="y", pady=20)
        tk.Frame(az, bg=BORDER, width=1).pack(side="left", fill="y", pady=30)
        az_right = tk.Frame(az, bg=PANEL)
        az_right.pack(side="left", fill="both", expand=True, pady=20)

        tk.Label(az_left, text="Adzuna API", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(0, 4))
        tk.Label(az_left, text="Free API — up to 250 requests/month.",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 4))
        az_portal_row = tk.Frame(az_left, bg=PANEL)
        az_portal_row.pack(anchor="w", padx=20, pady=(0, 10))
        tk.Label(az_portal_row, text="🔗 ", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        lnk_az = tk.Label(az_portal_row, text="developer.adzuna.com", bg=PANEL, fg=ACCENT2,
                           font=("Segoe UI", 9, "underline"), cursor="hand2")
        lnk_az.pack(side="left")
        lnk_az.bind("<Button-1>",
                    lambda e: __import__("webbrowser").open("https://developer.adzuna.com"))

        az_countries = (
            "Australia · Austria · Belgium · Brazil · Canada · France · Germany · India · Italy · "
            "Mexico · Netherlands · New Zealand · Poland · Singapore · South Africa · Spain · "
            "Switzerland · UK · USA"
        )
        tk.Label(az_left, text=f"Countries ({len(COUNTRIES)}):  {az_countries}",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9),
                 wraplength=400, justify="left").pack(anchor="w", padx=20, pady=(0, 16))

        az_fields = tk.Frame(az_left, bg=PANEL)
        az_fields.pack(anchor="w", padx=20, pady=(0, 16))

        tk.Label(az_fields, text="App ID", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.app_id_var = tk.StringVar(value=APP_ID)
        tk.Entry(az_fields, textvariable=self.app_id_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=28).grid(row=1, column=0, ipady=6, padx=(0, 20))

        tk.Label(az_fields, text="App Key", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=1, sticky="w")
        self.app_key_var = tk.StringVar(value=APP_KEY)
        tk.Entry(az_fields, textvariable=self.app_key_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=36).grid(row=1, column=1, ipady=6)

        az_btn = tk.Frame(az_left, bg=PANEL)
        az_btn.pack(anchor="w", padx=20)
        self._btn(az_btn, "Save Credentials", self._save_credentials, w=18).pack(side="left")

        self.cred_status = tk.Label(az_left, text="", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9))
        self.cred_status.pack(anchor="w", padx=20, pady=(8, 0))

        # Guide — right column
        tk.Label(az_right, text="How to get your free Adzuna credentials", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(0, 8))

        for num, text in [
            ("1", 'Go to developer.adzuna.com and click "Register"'),
            ("2", "Create a free account with your e-mail address"),
            ("3", 'After login, click "Create new application"'),
            ("4", 'Fill in any app name (e.g. "Job Search") and click "Save"'),
            ("5", "Copy the App ID and App Key shown on the dashboard"),
            ("6", 'Paste them into the fields above and click "Save Credentials"'),
        ]:
            make_step(az_right, num, text)


        # ── Reed sub-tab ──────────────────────────────────────────────────────
        rd = tk.Frame(tab_reed, bg=PANEL)
        rd.pack(fill="both", expand=True, padx=30, pady=20)

        rd_left = tk.Frame(rd, bg=PANEL)
        rd_left.pack(side="left", fill="y", pady=20)
        tk.Frame(rd, bg=BORDER, width=1).pack(side="left", fill="y", pady=30)
        rd_right = tk.Frame(rd, bg=PANEL)
        rd_right.pack(side="left", fill="both", expand=True, pady=20)

        tk.Label(rd_left, text="Reed API  (UK only)", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(0, 4))
        tk.Label(rd_left, text="Free API for UK job listings.",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 4))
        rd_portal_row = tk.Frame(rd_left, bg=PANEL)
        rd_portal_row.pack(anchor="w", padx=20, pady=(0, 10))
        tk.Label(rd_portal_row, text="🔗 ", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        lnk_rd_top = tk.Label(rd_portal_row, text="reed.co.uk/developers", bg=PANEL, fg=ACCENT2,
                               font=("Segoe UI", 9, "underline"), cursor="hand2")
        lnk_rd_top.pack(side="left")
        lnk_rd_top.bind("<Button-1>",
                        lambda e: __import__("webbrowser").open("https://www.reed.co.uk/developers"))
        tk.Label(rd_left, text="Countries (1):  United Kingdom",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 16))

        rd_fields = tk.Frame(rd_left, bg=PANEL)
        rd_fields.pack(anchor="w", padx=20, pady=(0, 16))

        tk.Label(rd_fields, text="API Key", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.reed_key_var = tk.StringVar(value=cfg.get("reed_key", ""))
        tk.Entry(rd_fields, textvariable=self.reed_key_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=52).grid(row=1, column=0, ipady=6)

        rd_btn = tk.Frame(rd_left, bg=PANEL)
        rd_btn.pack(anchor="w", padx=20)
        self._btn(rd_btn, "Save Credentials", self._save_credentials, w=18).pack(side="left")

        self.reed_status = tk.Label(rd_left, text="", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9))
        self.reed_status.pack(anchor="w", padx=20, pady=(8, 0))

        # Guide — right column
        tk.Label(rd_right, text="How to get your free Reed API key", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(0, 8))

        for num, text in [
            ("1", 'Go to reed.co.uk/developers and click "Register"'),
            ("2", "Create a free account with your e-mail address"),
            ("3", "Your API key is shown on the developer dashboard"),
            ("4", 'Paste it into the field above and click "Save Credentials"'),
            ("5", 'Enable the Reed (UK) checkbox in the Search tab'),
        ]:
            make_step(rd_right, num, text)


        # ── Findwork sub-tab ──────────────────────────────────────────────────
        fw = tk.Frame(tab_cj, bg=PANEL)
        fw.pack(fill="both", expand=True, padx=30, pady=20)

        fw_left = tk.Frame(fw, bg=PANEL)
        fw_left.pack(side="left", fill="y", pady=20)
        tk.Frame(fw, bg=BORDER, width=1).pack(side="left", fill="y", pady=30)
        fw_right = tk.Frame(fw, bg=PANEL)
        fw_right.pack(side="left", fill="both", expand=True, pady=20)

        tk.Label(fw_left, text="Findwork API", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(0, 4))
        tk.Label(fw_left, text="Tech & remote job listings worldwide. Free API key.",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 4))
        fw_portal_row = tk.Frame(fw_left, bg=PANEL)
        fw_portal_row.pack(anchor="w", padx=20, pady=(0, 10))
        tk.Label(fw_portal_row, text="🔗 ", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        lnk_fw_top = tk.Label(fw_portal_row, text="findwork.dev/developers", bg=PANEL, fg=ACCENT2,
                               font=("Segoe UI", 9, "underline"), cursor="hand2")
        lnk_fw_top.pack(side="left")
        lnk_fw_top.bind("<Button-1>",
                        lambda e: __import__("webbrowser").open("https://findwork.dev/developers/"))
        tk.Label(fw_left, text="Coverage:  Tech & remote jobs globally — aggregated from HN, RemoteOK, WeWorkRemotely & more.",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9),
                 wraplength=400, justify="left").pack(anchor="w", padx=20, pady=(0, 16))

        fw_fields = tk.Frame(fw_left, bg=PANEL)
        fw_fields.pack(anchor="w", padx=20, pady=(0, 16))
        tk.Label(fw_fields, text="API Key", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.findwork_key_var = tk.StringVar(value=cfg.get("findwork_key", ""))
        tk.Entry(fw_fields, textvariable=self.findwork_key_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=52).grid(row=1, column=0, ipady=6)

        fw_btn = tk.Frame(fw_left, bg=PANEL)
        fw_btn.pack(anchor="w", padx=20)
        self._btn(fw_btn, "Save Credentials", self._save_credentials, w=18).pack(side="left")
        self.fw_status = tk.Label(fw_left, text="", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9))
        self.fw_status.pack(anchor="w", padx=20, pady=(8, 0))

        # Guide — right column
        tk.Label(fw_right, text="How to get your free Findwork API key", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(0, 8))
        for num, text in [
            ("1", "Go to findwork.dev and click \"Sign Up\" (top right)"),
            ("2", "Register with your e-mail address — completely free"),
            ("3", "After login, go to findwork.dev/developers"),
            ("4", "Your API key is shown on that page — copy it"),
            ("5", "Paste it above and click \"Save Credentials\""),
        ]:
            make_step(fw_right, num, text)

        # ── Jooble sub-tab ────────────────────────────────────────────────────
        jb = tk.Frame(tab_jooble, bg=PANEL)
        jb.pack(fill="both", expand=True, padx=30, pady=20)

        jb_left = tk.Frame(jb, bg=PANEL)
        jb_left.pack(side="left", fill="y", pady=20)
        tk.Frame(jb, bg=BORDER, width=1).pack(side="left", fill="y", pady=30)
        jb_right = tk.Frame(jb, bg=PANEL)
        jb_right.pack(side="left", fill="both", expand=True, pady=20)

        tk.Label(jb_left, text="Jooble API", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(0, 4))
        tk.Label(jb_left, text="Global job aggregator. Free API key via e-mail request.",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 4))
        jb_portal_row = tk.Frame(jb_left, bg=PANEL)
        jb_portal_row.pack(anchor="w", padx=20, pady=(0, 10))
        tk.Label(jb_portal_row, text="🔗 ", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        lnk_jb_top = tk.Label(jb_portal_row, text="jooble.org/api/about", bg=PANEL, fg=ACCENT2,
                               font=("Segoe UI", 9, "underline"), cursor="hand2")
        lnk_jb_top.pack(side="left")
        lnk_jb_top.bind("<Button-1>",
                        lambda e: __import__("webbrowser").open("https://jooble.org/api/about"))
        jb_countries = (
            "Argentina · Australia · Austria · Belgium · Brazil · Canada · Chile · "
            "Colombia · Czech Republic · Denmark · Finland · France · Germany · Greece · "
            "Hungary · India · Ireland · Italy · Mexico · Netherlands · New Zealand · "
            "Norway · Peru · Poland · Portugal · Romania · Russia · South Africa · Spain · "
            "Sweden · Switzerland · UK · Ukraine · USA · Venezuela · and more..."
        )
        tk.Label(jb_left, text=f"Countries (70+):  {jb_countries}",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9),
                 wraplength=400, justify="left").pack(anchor="w", padx=20, pady=(0, 16))

        jb_fields = tk.Frame(jb_left, bg=PANEL)
        jb_fields.pack(anchor="w", padx=20, pady=(0, 16))
        tk.Label(jb_fields, text="API Key", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.jooble_key_var = tk.StringVar(value=cfg.get("jooble_key", ""))
        tk.Entry(jb_fields, textvariable=self.jooble_key_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=52).grid(row=1, column=0, ipady=6)

        jb_btn = tk.Frame(jb_left, bg=PANEL)
        jb_btn.pack(anchor="w", padx=20)
        self._btn(jb_btn, "Save Credentials", self._save_credentials, w=18).pack(side="left")
        self.jooble_status = tk.Label(jb_left, text="", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9))
        self.jooble_status.pack(anchor="w", padx=20, pady=(8, 0))

        # Guide — right column
        tk.Label(jb_right, text="How to get your Jooble API key", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(0, 8))
        for num, text in [
            ("1", "Go to jooble.org/api/about and fill in the registration form"),
            ("2", "Submit the form — Jooble will send your API key by e-mail"),
            ("3", "Check your inbox (and spam folder) for an e-mail from Jooble"),
            ("4", "Copy the API key from the e-mail"),
            ("5", "Paste it above and click \"Save Credentials\""),
        ]:
            make_step(jb_right, num, text)



        # ── HeadHunter ────────────────────────────────────────────────────────
        hh = tk.Frame(tab_hh, bg=PANEL)
        hh.pack(fill="both", expand=True, padx=30, pady=20)

        hh_left = tk.Frame(hh, bg=PANEL)
        hh_left.pack(side="left", fill="y", pady=20)
        tk.Frame(hh, bg=BORDER, width=1).pack(side="left", fill="y", pady=30)
        hh_right = tk.Frame(hh, bg=PANEL)
        hh_right.pack(side="left", fill="both", expand=True, pady=20)

        tk.Label(hh_left, text="HeadHunter API  (Russia / CIS)", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(0, 4))
        tk.Label(hh_left, text="Russia's largest job platform. Requires OAuth token (free registration).",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 4))
        hh_portal_row = tk.Frame(hh_left, bg=PANEL)
        hh_portal_row.pack(anchor="w", padx=20, pady=(0, 10))
        tk.Label(hh_portal_row, text="🔗 ", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="left")
        lnk_hh_top = tk.Label(hh_portal_row, text="dev.hh.ru", bg=PANEL, fg=ACCENT2,
                               font=("Segoe UI", 9, "underline"), cursor="hand2")
        lnk_hh_top.pack(side="left")
        lnk_hh_top.bind("<Button-1>",
                        lambda e: __import__("webbrowser").open("https://dev.hh.ru"))
        tk.Label(hh_left, text="Countries:  Russia 🇷🇺  Kazakhstan 🇰🇿  Belarus 🇧🇾  Ukraine 🇺🇦  and more CIS",
                 bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 16))

        hh_fields = tk.Frame(hh_left, bg=PANEL)
        hh_fields.pack(anchor="w", padx=20, pady=(0, 16))
        tk.Label(hh_fields, text="Access Token", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w")
        self.hh_token_var = tk.StringVar(value=cfg.get("hh_token", ""))
        tk.Entry(hh_fields, textvariable=self.hh_token_var, bg=BG, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=("Segoe UI", 11),
                 width=52).grid(row=1, column=0, ipady=6)

        hh_btn = tk.Frame(hh_left, bg=PANEL)
        hh_btn.pack(anchor="w", padx=20)
        self._btn(hh_btn, "Save Credentials", self._save_credentials, w=18).pack(side="left")
        self.hh_status = tk.Label(hh_left, text="", bg=PANEL, fg=SUCCESS, font=("Segoe UI", 9))
        self.hh_status.pack(anchor="w", padx=20, pady=(8, 0))

        # Guide — right column
        tk.Label(hh_right, text="How to get your HeadHunter token", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=20, pady=(0, 8))
        for num, text in [
            ("1", "Go to dev.hh.ru and log in with your hh.ru account"),
            ("2", "Click \"Создать приложение\" (Create application)"),
            ("3", "Fill in the app name and redirect URI (e.g. https://localhost)"),
            ("4", "Open this URL in your browser (replace YOUR_CLIENT_ID):"),
            ("4b", "https://hh.ru/oauth/authorize?response_type=code&client_id=YOUR_CLIENT_ID"),
            ("5", "Authorize and copy the 'code' from the redirect URL"),
            ("6", "Exchange the code for a token via the hh.ru token endpoint"),
            ("7", "Paste the access_token above and click \"Save Credentials\""),
        ]:
            make_step(hh_right, num, text)


        # Show first tab initially
        _cred_switch(0)

    def _save_provider_state(self):
        _save_config({"providers": {
            "adzuna":     self.use_adzuna.get(),
            "reed":       self.use_reed.get(),
            "findwork":   self.use_findwork.get(),
            "arbeitnow":  self.use_arbeitnow.get(),
            "remoteok":   self.use_remoteok.get(),
            "jooble":     self.use_jooble.get(),
            "themuse":    self.use_themuse.get(),
            "bundesag":   self.use_bundesag.get(),
            "headhunter": self.use_headhunter.get(),
            "wwr":        self.use_wwr.get(),
            "remotive":   self.use_remotive.get(),
            "himalayas":  self.use_himalayas.get(),
        }})

    def _save_credentials(self):
        global APP_ID, APP_KEY
        new_id   = self.app_id_var.get().strip()
        new_key  = self.app_key_var.get().strip()
        if not new_id or not new_key:
            self.cred_status.config(text="App ID and App Key cannot be empty.", fg=DANGER)
            return
        APP_ID  = new_id
        APP_KEY = new_key
        try:
            _save_config({
                "app_id":       new_id,
                "app_key":      new_key,
                "reed_key":     self.reed_key_var.get().strip(),
                "findwork_key": self.findwork_key_var.get().strip(),
                "jooble_key":   self.jooble_key_var.get().strip(),
                "hh_token":     self.hh_token_var.get().strip(),
            })
            msg = "Credentials saved."
            for lbl in (self.cred_status, self.reed_status,
                        self.fw_status, self.jooble_status,
                        self.hh_status):
                lbl.config(text=msg, fg=SUCCESS)
        except Exception as e:
            err = f"Session only (file error: {e})"
            for lbl in (self.cred_status, self.reed_status,
                        self.fw_status, self.jooble_status,
                        self.hh_status):
                lbl.config(text=err, fg=ACCENT2)

    # ── Help / About popups ───────────────────────────────────────────────────
    def _show_help(self):
        win = tk.Toplevel(self)
        win.title("Help — J🔍B Search Tool")
        win.geometry("700x600")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        self._help_content(win)
        win.bind("<Destroy>", lambda e: win.unbind_all("<MouseWheel>"))

    def _show_about(self):
        win = tk.Toplevel(self)
        win.title("About — J🔍B Search Tool")
        win.geometry("500x440")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()
        self._about_content(win)

    # ── Help content (shared by popup) ────────────────────────────────────────
    def _help_content(self, parent):
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=BG)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        P = 30

        def section(title):
            tk.Frame(inner, bg=BORDER, height=1).pack(fill="x", padx=P, pady=(24, 0))
            tk.Label(inner, text=title, bg=BG, fg=ACCENT2,
                     font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=P, pady=(10, 6))

        def step(num, text):
            row = tk.Frame(inner, bg=BG)
            row.pack(anchor="w", padx=P, pady=3)
            tk.Label(row, text=str(num), bg=ACCENT, fg="white",
                     font=("Segoe UI", 8, "bold"), width=2,
                     relief="flat").pack(side="left", padx=(0, 12))
            tk.Label(row, text=text, bg=BG, fg=TEXT,
                     font=("Segoe UI", 10), justify="left").pack(side="left")

        def info(label, text):
            row = tk.Frame(inner, bg=BG)
            row.pack(anchor="w", padx=P, pady=3)
            tk.Label(row, text=label, bg=BG, fg=ACCENT2,
                     font=("Segoe UI", 10, "bold"), width=18,
                     anchor="w").pack(side="left")
            tk.Label(row, text=text, bg=BG, fg=TEXT,
                     font=("Segoe UI", 10), justify="left").pack(side="left")

        def faq(question, answer):
            tk.Label(inner, text=question, bg=BG, fg=TEXT,
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(
                         anchor="w", padx=P, pady=(10, 1))
            tk.Label(inner, text=answer, bg=BG, fg=SUBTEXT,
                     font=("Segoe UI", 10), anchor="w", justify="left").pack(
                         anchor="w", padx=P + 12)

        logo_row = tk.Frame(inner, bg=BG)
        logo_row.pack(anchor="w", padx=P, pady=(24, 2))
        self._make_logo(logo_row, 16, BG).pack(side="left")
        tk.Label(logo_row, text=" Search Tool", bg=BG, fg=TEXT,
                 font=("Segoe UI", 16, "bold")).pack(side="left")
        tk.Label(inner, text="Search, save and analyse job listings from 12 portals — automatically, every day.",
                 bg=BG, fg=SUBTEXT, font=("Segoe UI", 10)).pack(anchor="w", padx=P)

        section("Quick Start  —  no API key required")
        step(1, 'Open the Search tab. Leave Job Title empty to browse all jobs, or type e.g. "Data Analyst".')
        step(2, 'Select a Country and tick one or more providers in the "No key" row (e.g. Arbeitnow, RemoteOK).')
        step(3, 'Set how many Results you want and click Search.')
        step(4, 'Results appear in the table. Double-click any row to open the listing in your browser.')
        step(5, 'Click Save to export results to Excel. The file appears in Saved Results automatically.')

        section("Providers")
        info("No key needed",  "BA, Arbeitnow, The Muse, RemoteOK, WWR (WeWorkRemotely), Remotive, Himalayas")
        info("API key needed", "Adzuna, Reed, Findwork, Jooble, HeadHunter (OAuth token)")
        info("Country filter", "Selecting a country automatically enables only the providers available\n"
                               "for that country. Greyed-out chips are not supported for the selection.")
        info("Keyword search", 'All words must match — "data analyst" finds jobs containing both\n'
                               '"data" AND "analyst" anywhere in the title or description.')
        info("Results limit",  "Each provider is queried until it reaches the requested count or\n"
                               "runs out of results. RSS-based providers (WWR, Remotive) are limited\n"
                               "to ~30–100 jobs by the feed itself. Himalayas returns up to 500+ remote jobs.")

        section("Tab Guide")
        info("Search",        "Live job search. Sort results by clicking any column header (▲ / ▼).\n"
                              "Double-click a row to open the URL. Click Save to export to Excel.")
        info("Saved Results", "Browse saved Excel files. Click a file on the left to load its jobs.\n"
                              "Sort by any column. Double-click a row to open the listing.")
        info("Auto Run",      "Schedule a daily automated search via Windows Task Scheduler.\n"
                              "New jobs are appended; duplicates are skipped. Sort the run list by column.")
        info("Analytics",     "Charts and KPI cards based on your saved Excel files.\n"
                              "Select a file, click Refresh. Export a PDF report with 'Export PDF'.")
        info("Credentials",   "Enter and save API keys for paid providers. Stored in config.json.")

        section("Setting up API Keys")
        info("Adzuna",      "Register at adzuna.com/developers — free tier: up to 250 requests/month.\n"
                            "Enter App ID and App Key in the Credentials tab.")
        info("Reed",        "Register at reed.co.uk/developers — free API key, UK jobs only.")
        info("Findwork",    "Register at findwork.dev — free API key, tech & remote jobs worldwide.")
        info("Jooble",      "Register at jooble.org/api/about — free API key, global jobs.")
        info("HeadHunter",  "Create an app at dev.hh.ru, get an OAuth token. Russia & CIS countries.")

        section("Setting up Auto Run")
        step(1, "Run a search with the Job Title and Location you want to track daily.")
        step(2, 'Go to Auto Run → click "Add from Search Tab" to add the current search.')
        step(3, 'Set the time (e.g. "08:00") and click "Schedule Daily".')
        step(4, "Windows Task Scheduler runs the search every day at that time.")
        step(5, "New jobs are appended to the Excel file. Duplicates are skipped automatically.")
        step(6, 'To stop: select the entry and click "Remove Selected".')

        section("Analytics")
        info("Load data",    'Select a saved file from the dropdown and click "Refresh".')
        info("KPI cards",    "Shows total jobs, unique companies, average salary and date range at a glance.")
        info("Charts",       "Top Companies, Salary Distribution, Jobs by Location, Salary by Location,\n"
                             "Monthly Activity, Salary Trend, Top Keywords.")
        info("Export PDF",   'Click "Export PDF" to save all charts and stats as a PDF report.')

        section("FAQ")
        faq("Are my API keys safe?",
            "Yes. All keys are stored only in a local config file on your device.\n"
            "The app never sends your keys anywhere except directly to the respective API provider.\n"
            "This is called BYOK (Bring Your Own Key) — your data stays yours.")
        faq("Where are my Excel files saved?",
            "In a  jobs/  folder next to the app, organised by search term.\n"
            "Example:  jobs/Data_Analyst_Berlin/Data_Analyst_Berlin.xlsx")
        faq("Why does a provider return 0 jobs?",
            "Possible reasons: the keyword doesn't match any current listings,\n"
            "the API is temporarily unavailable, or a required API key is missing.")
        faq("Why is the provider chip greyed out?",
            "That provider doesn't cover the selected country. Change the country\n"
            "or switch to a global provider (Findwork, Jooble, RemoteOK, WWR, Remotive, Himalayas).")
        faq("The scheduled run found 0 jobs — what happened?",
            "Check job_search.log in the app folder for details.\n"
            "Most common cause: invalid credentials or no internet connection.")
        faq("How do I sort the results table?",
            "Click any column header to sort ascending (▲).\n"
            "Click again to sort descending (▼). Works in Search, Saved Results and Auto Run.")
        faq("How do I update the scheduled search?",
            "Remove the old entry, run a new search with the updated settings,\n"
            'then use "Add from Search Tab" and reschedule.')

        tk.Frame(inner, bg=BG, height=30).pack()

    # ── About content (shared by popup) ───────────────────────────────────────
    def _about_content(self, parent):
        outer = tk.Frame(parent, bg=BG)
        outer.pack(fill="both", expand=True)

        center = tk.Frame(outer, bg=BG)
        center.place(relx=0.5, rely=0.5, anchor="center")

        def _link_btn(parent, text, url, color):
            lbl = tk.Label(parent, text=text, bg=color, fg="white",
                           font=("Segoe UI", 9, "bold"),
                           padx=14, pady=7, relief="flat", cursor="hand2")
            lbl.pack(side="left", padx=5)
            lbl.bind("<Button-1>", lambda e: __import__("webbrowser").open(url))

        # ── Icon + title ──────────────────────────────────────────────────────
        self._make_logo(center, 42, BG).pack(pady=(0, 4))
        tk.Label(center, text="Search Tool", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 11)).pack(pady=(0, 4))

        # Version badge
        badge = tk.Frame(center, bg=ACCENT, padx=10, pady=3)
        badge.pack(pady=(6, 24))
        tk.Label(badge, text=f"v {APP_VERSION}", bg=ACCENT, fg="white",
                 font=("Segoe UI", 8, "bold")).pack()

        tk.Frame(center, bg=BORDER, height=1, width=440).pack(pady=(0, 20))

        # ── Author ────────────────────────────────────────────────────────────
        tk.Label(center, text="Built by", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack()
        name_lbl = tk.Label(center, text="Todor Vankov", bg=BG, fg=TEXT,
                            font=("Segoe UI", 15, "bold"), cursor="hand2")
        name_lbl.pack(pady=(3, 12))
        name_lbl.bind("<Button-1>",
                      lambda e: __import__("webbrowser").open("https://www.todorvankov.com"))

        links_row = tk.Frame(center, bg=BG)
        links_row.pack(pady=(0, 24))
        _link_btn(links_row, "🌐  Website",  "https://www.todorvankov.com",                                "#0ea5e9")
        _link_btn(links_row, "🔗  LinkedIn", "https://www.linkedin.com/in/todor-dimitrov-vankov-1398bb2b/", "#0a66c2")
        _link_btn(links_row, "🐙  GitHub",   "https://github.com/tvankov",                                 "#334155")

        tk.Frame(center, bg=BORDER, height=1, width=440).pack(pady=(0, 20))

        # ── Tech stack ────────────────────────────────────────────────────────
        tk.Label(center, text="Built with", bg=BG, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(pady=(0, 10))

        stack_row = tk.Frame(center, bg=BG)
        stack_row.pack(pady=(0, 20))
        for tech, color in [("Python",     "#3b82f6"),
                             ("Tkinter",    "#8b5cf6"),
                             ("openpyxl",   "#10b981"),
                             ("matplotlib", "#f59e0b"),
                             ("12 APIs",    "#ef4444")]:
            tk.Label(stack_row, text=tech, bg=color, fg="white",
                     font=("Segoe UI", 8, "bold"),
                     padx=9, pady=4, relief="flat").pack(side="left", padx=3)

        tk.Frame(center, bg=BORDER, height=1, width=440).pack(pady=(0, 14))

        tk.Label(center,
                 text="salary analytics  ·  daily auto-run  ·  PDF reports",
                 bg=BG, fg=SUBTEXT, font=("Segoe UI", 8)).pack()

    # ── Update check ──────────────────────────────────────────────────────────
    _VERSION_URL = ("https://raw.githubusercontent.com/"
                    "tvankov/job-search-tool/main/version.json")

    def _check_for_updates(self):
        def _fetch():
            try:
                resp = requests.get(self._VERSION_URL, timeout=6)
                if resp.status_code != 200:
                    return
                data       = resp.json()
                latest     = data.get("version", "0")
                dl_url     = data.get("download_url", "")
                changelog  = data.get("changelog", "")
                if self._is_newer(latest, APP_VERSION):
                    self.after(0, lambda: self._show_update_banner(
                        latest, dl_url, changelog))
            except Exception:
                pass
        threading.Thread(target=_fetch, daemon=True).start()

    @staticmethod
    def _is_newer(remote: str, local: str) -> bool:
        try:
            r = tuple(int(x) for x in remote.strip().split("."))
            l = tuple(int(x) for x in local.strip().split("."))
            return r > l
        except Exception:
            return False

    def _show_update_banner(self, version, url, changelog):
        if self._update_banner:
            return
        bar = tk.Frame(self, bg="#0c4a6e", pady=6)
        bar.pack(fill="x", after=self.winfo_children()[0])
        self._update_banner = bar

        msg = f"🚀  Update available  —  v {version} is ready"
        if changelog:
            msg += f"   ·   {changelog}"
        tk.Label(bar, text=msg, bg="#0c4a6e", fg="#bae6fd",
                 font=("Segoe UI", 9)).pack(side="left", padx=16)

        def _open():
            __import__("webbrowser").open(url)

        def _dismiss():
            bar.destroy()
            self._update_banner = None

        self._btn(bar, "⬇  Download", _open,
                  color=ACCENT, w=11).pack(side="right", padx=(0, 8))
        self._btn(bar, "✕", _dismiss,
                  color="#0c4a6e", w=3).pack(side="right", padx=(0, 4))

    # ── Footer ────────────────────────────────────────────────────────────────
    def _footer(self):
        bar = tk.Frame(self, bg=PANEL, pady=8)
        bar.pack(fill="x", side="bottom")

        self.status_lbl = tk.Label(bar, text="Ready — enter a search and click 🔍 Search",
                                   bg=PANEL, fg=SUBTEXT, font=("Segoe UI", 9))
        self.status_lbl.pack(side="left", padx=16)

        self._btn(bar, "🐛  Report Bug", self._show_bug_report,
                  color="#1e3a4a", w=12).pack(side="right", padx=(0, 8))

        self._clear_btn = self._btn(bar, "🗑  Clear Results", self._clear, color="#334155", w=13)
        self._clear_btn.pack(side="right", padx=(22, 4))

    # ── Bug Report ────────────────────────────────────────────────────────────
    def _show_bug_report(self):
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "error.log")

        win = tk.Toplevel(self)
        win.title("Report a Bug")
        win.geometry("680x530")
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()

        # Header
        hdr = tk.Frame(win, bg=PANEL, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🐛  Report a Bug", bg=PANEL, fg=ACCENT2,
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=20)
        tk.Label(hdr, text=f"v {APP_VERSION}", bg=PANEL, fg=SUBTEXT,
                 font=("Segoe UI", 9)).pack(side="right", padx=20)

        # Description
        desc = tk.Frame(win, bg=BG)
        desc.pack(fill="x", padx=20, pady=(14, 6))
        tk.Label(desc,
                 text="Copy the error log below and send it to the developer.\n"
                      "The log contains only technical error info — no personal data.",
                 bg=BG, fg=SUBTEXT, font=("Segoe UI", 9),
                 justify="left").pack(anchor="w")

        tk.Frame(win, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(6, 10))

        # Log content
        log_frame = tk.Frame(win, bg=BG, bd=0, highlightthickness=0)
        log_frame.pack(fill="both", expand=True, padx=20)

        log_sb = ttk.Scrollbar(log_frame, orient="vertical")
        log_sb.pack(side="right", fill="y")

        log_text = tk.Text(log_frame, bg=PANEL, fg=TEXT,
                           font=("Consolas", 9), relief="flat",
                           borderwidth=0, highlightthickness=0,
                           wrap="none", yscrollcommand=log_sb.set,
                           state="normal")
        log_text.pack(fill="both", expand=True)
        log_sb.config(command=log_text.yview)

        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
            log_text.insert("1.0", content if content else "No errors logged yet.")
        else:
            log_text.insert("1.0", "No errors logged yet. The log file will be created automatically\n"
                                   "when an error occurs.")
        log_text.config(state="disabled")
        log_text.see("end")

        # Buttons
        tk.Frame(win, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(10, 0))
        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=20, pady=10)

        def _copy():
            win.clipboard_clear()
            win.clipboard_append(log_text.get("1.0", "end").strip())
            copy_btn.config(text="✓  Copied!")
            win.after(2000, lambda: copy_btn.config(text="📋  Copy Log"))

        def _email():
            import urllib.parse
            body = urllib.parse.quote(log_text.get("1.0", "end").strip()[:1800])
            subject = urllib.parse.quote(f"Bug Report — Job Search Tool v{APP_VERSION}")
            __import__("webbrowser").open(
                f"mailto:todor@todorvankov.com?subject={subject}&body={body}")

        def _clear_log():
            if messagebox.askyesno("Clear Log", "Delete all entries in the error log?",
                                   parent=win):
                open(log_path, "w").close()
                log_text.config(state="normal")
                log_text.delete("1.0", "end")
                log_text.insert("1.0", "Log cleared.")
                log_text.config(state="disabled")

        copy_btn = self._btn(btn_row, "📋  Copy Log",  _copy,  color=ACCENT,    w=13)
        copy_btn.pack(side="left", padx=(0, 8))
        self._btn(btn_row, "📧  Send Email", _email, color="#0f766e", w=13).pack(side="left", padx=(0, 8))
        self._btn(btn_row, "🗑  Clear Log",  _clear_log, color="#7f1d1d", w=12).pack(side="left")
        self._btn(btn_row, "Close",          win.destroy, color="#334155", w=8).pack(side="right")

    # ── Helper ────────────────────────────────────────────────────────────────
    def _bind_combobox_typeahead(self, cb, values):
        typed = {"buf": "", "after": None}
        def on_key(e):
            ch = e.char.lower()
            if not ch or not ch.isprintable():
                return
            if typed["after"]:
                cb.after_cancel(typed["after"])
            typed["buf"] += ch
            prefix = typed["buf"]
            match = next((v for v in values if v.lower().startswith(prefix)), None)
            if match:
                cb.set(match)
            typed["after"] = cb.after(800, lambda: typed.update(buf="", after=None))
        cb.bind("<KeyPress>", on_key)

    def _btn(self, parent, text, cmd, color=ACCENT, w=None):
        kw = dict(text=text, command=cmd, bg=color, fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat",
                  cursor="hand2", activebackground=ACCENT2,
                  activeforeground="white", pady=6, padx=12)
        if w:
            kw["width"] = w
        b = tk.Button(parent, **kw)
        b.bind("<Enter>", lambda e: b.config(bg=ACCENT2))
        b.bind("<Leave>", lambda e: b.config(bg=color))
        return b

    def _on_close(self):
        self._analytics_cancel = True
        plt.close("all")
        self.quit()
        self.destroy()
        import sys; sys.exit(0)

    def _set_status(self, msg, ok=True):
        self.status_lbl.config(text=msg, fg=SUCCESS if ok else DANGER)

    # ── Tree sorting ──────────────────────────────────────────────────────────
    def _sort_search_tree(self, col):
        if not self.jobs:
            return
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = False

        _job_key = {
            "title":    "title",
            "company":  "company",
            "location": "location",
            "salary":   "salary_min",
            "date":     "created",
            "source":   "source",
            "url":      "url",
        }
        key = _job_key[col]

        def _sort_val(job):
            v = job.get(key, "") or ""
            try:
                return (0, float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                return (1, str(v).lower())

        self.jobs.sort(key=_sort_val, reverse=self._sort_rev)

        # Update headings: show arrow only on active column
        for c, lbl in self._tree_col_labels.items():
            if c == col:
                arrow = " ▼" if self._sort_rev else " ▲"
                self.tree.heading(c, text=lbl + arrow)
            else:
                self.tree.heading(c, text=lbl)

        # Repopulate tree
        for row in self.tree.get_children():
            self.tree.delete(row)
        for i, job in enumerate(self.jobs):
            s_min = job.get("salary_min")
            s_max = job.get("salary_max")
            try:
                salary = (f"{int(float(s_min)):,} – {int(float(s_max)):,}" if s_min and s_max
                          else f"from {int(float(s_min)):,}" if s_min else "")
            except Exception:
                salary = ""
            self.tree.insert("", "end", iid=str(i),
                             tags=("odd" if i % 2 else "even",),
                             values=(job["title"], job["company"], job["location"],
                                     salary, job["created"], job["source"], job["url"]))

    def _sort_treeview(self, tree, col, col_labels, state):
        if state["col"] == col:
            state["rev"] = not state["rev"]
        else:
            state["col"] = col
            state["rev"] = False

        col_names = list(col_labels.keys())
        col_idx   = col_names.index(col)

        rows = [tree.item(iid)["values"] for iid in tree.get_children()]

        def _key(vals):
            v = vals[col_idx] if col_idx < len(vals) else ""
            try:
                return (0, float(str(v).replace(",", "").replace(" ", "")))
            except Exception:
                return (1, str(v).lower())

        rows.sort(key=_key, reverse=state["rev"])

        for c, lbl in col_labels.items():
            arrow = (" ▼" if state["rev"] else " ▲") if c == col else ""
            tree.heading(c, text=lbl + arrow)

        for iid in tree.get_children():
            tree.delete(iid)
        for i, vals in enumerate(rows):
            tree.insert("", "end", iid=str(i),
                        tags=("odd" if i % 2 else "even",),
                        values=vals)

    # ── Search ────────────────────────────────────────────────────────────────
    def _search(self):
        what    = self.what_var.get().strip()
        where   = self.where_var.get().strip()
        country = COUNTRIES.get(self.country_var.get(), "de")
        results = self.results_var.get()
        sort_by = "relevance"

        active = [v for v in [
            self.use_adzuna, self.use_reed, self.use_findwork,
            self.use_arbeitnow, self.use_remoteok, self.use_jooble,
            self.use_themuse, self.use_bundesag,
            self.use_headhunter, self.use_wwr, self.use_remotive,
            self.use_himalayas,
        ] if v.get()]
        if not active:
            self._set_status("Select at least one provider.", ok=False)
            return

        per_provider = results

        self._empty_state.place_forget()
        self._set_status("Searching…")
        self.update_idletasks()

        all_jobs     = []
        seen_urls    = set()
        auth_errors  = []
        cfg          = _load_config()

        def _add(jobs):
            for j in jobs:
                if len(all_jobs) >= results:
                    break
                if j["url"]:
                    if j["url"] in seen_urls:
                        continue
                    seen_urls.add(j["url"])
                all_jobs.append(j)

        def _fetch(name, fn):
            if len(all_jobs) >= results:
                return
            self._set_status(f"Searching {name}…  ({len(all_jobs)} so far)")
            self.update_idletasks()
            try:
                _add(fn())
            except AuthError:
                auth_errors.append(name)

        try:
            if self.use_adzuna.get():
                _fetch("Adzuna", lambda: AdzunaProvider(
                    self.app_id_var.get().strip(),
                    self.app_key_var.get().strip(),
                ).search(what, where, country=country, results=per_provider,
                         sort_by=sort_by,
                         salary_min=self.salary_min_var.get().strip() or None,
                         salary_max=self.salary_max_var.get().strip() or None,
                         full_time=self.fulltime_var.get(),
                         permanent=self.permanent_var.get()))

            if self.use_reed.get():
                _fetch("Reed", lambda: ReedProvider(
                    self.reed_key_var.get().strip()).search(what, where, results=per_provider))

            if self.use_findwork.get():
                _fetch("Findwork", lambda: FindworkProvider(
                    cfg.get("findwork_key", "")).search(what, where, results=per_provider))

            if self.use_jooble.get():
                _fetch("Jooble", lambda: JoobleProvider(
                    cfg.get("jooble_key", "")).search(what, where, results=per_provider))

            if self.use_arbeitnow.get():
                _fetch("Arbeitnow", lambda: ArbeitnowProvider().search(what, where, results=per_provider))

            if self.use_bundesag.get():
                _fetch("Bundesagentur", lambda: BundesagenturProvider().search(what, where, results=per_provider))

            if self.use_remoteok.get():
                _fetch("RemoteOK", lambda: RemoteOKProvider().search(what, where, results=per_provider))

            if self.use_themuse.get():
                _fetch("The Muse", lambda: TheMuseProvider(
                    cfg.get("themuse_key", "")).search(what, where, results=per_provider))


            if self.use_headhunter.get():
                _fetch("HeadHunter", lambda: HeadHunterProvider(
                    cfg.get("hh_token", "")).search(what, where, results=per_provider))

            if self.use_wwr.get():
                _fetch("WWR", lambda: WeWorkRemotelyProvider().search(what, where, results=per_provider))

            if self.use_remotive.get():
                _fetch("Remotive", lambda: RemotiveProvider().search(what, where, results=per_provider))

            if self.use_himalayas.get():
                _fetch("Himalayas", lambda: HimalayasProvider().search(what, where, results=per_provider))

            self.jobs = all_jobs[:results]
            self._sort_col = None
            self._sort_rev = False
            for c, lbl in self._tree_col_labels.items():
                self.tree.heading(c, text=lbl)
            for row in self.tree.get_children():
                self.tree.delete(row)

            self._set_status(f"Loading {len(self.jobs)} jobs…")
            self.update_idletasks()
            for i, job in enumerate(self.jobs):
                s_min = job.get("salary_min")
                s_max = job.get("salary_max")
                try:
                    salary = (f"{int(float(s_min)):,} – {int(float(s_max)):,}" if s_min and s_max
                              else f"from {int(float(s_min)):,}" if s_min else "")
                except Exception:
                    salary = ""
                self.tree.insert("", "end", iid=str(i),
                                 tags=("odd" if i % 2 else "even",),
                                 values=(job["title"], job["company"], job["location"],
                                         salary, job["created"], job["source"], job["url"]))
                if i % 50 == 49:
                    self.update_idletasks()

            if not self.jobs:
                self._empty_state.place(relx=0.5, rely=0.5, anchor="center")
                self._set_status("No jobs found — try different keywords or providers", ok=False)
            else:
                hint = "  (no keyword — showing latest)" if not what else ""
                self._set_status(f"✓ {len(self.jobs)} jobs found{hint}")

            if auth_errors:
                names = ", ".join(auth_errors)
                self._set_status(
                    f"⚠  Invalid API key: {names} — check Credentials tab", ok=False)

        except requests.exceptions.ConnectionError:
            self._empty_state.place(relx=0.5, rely=0.5, anchor="center")
            self._set_status("No internet connection", ok=False)
        except Exception as e:
            self._empty_state.place(relx=0.5, rely=0.5, anchor="center")
            self._set_status(f"Error: {e}", ok=False)

    # ── Double click → open URL ───────────────────────────────────────────────
    def _open_link(self, event):
        sel = self.tree.selection()
        if sel:
            url = self.tree.item(sel[0])["values"][6]
            if url:
                __import__("webbrowser").open(url)

    # ── Clear ─────────────────────────────────────────────────────────────────
    def _clear(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.jobs = []
        self._empty_state.place(relx=0.5, rely=0.5, anchor="center")
        self._set_status("Results cleared")

    # ── Export to Excel ───────────────────────────────────────────────────────
    def _export_excel(self, silent=False):
        if not self.jobs:
            if not silent:
                messagebox.showwarning("No data", "Please run a search first.")
            return

        slug     = (self.what_var.get().strip() + "_" + self.where_var.get().strip()).replace(" ", "_")
        base_dir = os.path.join(_app_dir(), "jobs", slug)
        os.makedirs(base_dir, exist_ok=True)
        path = os.path.join(base_dir, f"{slug}.xlsx")

        headers     = ["Title", "Company", "Location", "Salary Min (€)", "Salary Max (€)",
                       "Posted", "Link", "Description", "Source", "Imported On"]
        hdr_fill    = PatternFill("solid", start_color="0ea5e9")
        hdr_font    = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hdr_align   = Alignment(horizontal="center", vertical="center")
        thin_border = Border(bottom=Side(style="thin", color="334155"))
        odd_fill    = PatternFill("solid", start_color="1e293b")
        even_fill   = PatternFill("solid", start_color="162032")
        data_font   = Font(name="Arial", size=10, color="F1F5F9")
        link_font   = Font(name="Arial", size=10, color="38bdf8", underline="single")

        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb["Job Results"]
            existing_urls = {str(r[6]) for r in ws.iter_rows(min_row=2, values_only=True) if r[6]}
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Job Results"
            existing_urls = set()
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = hdr_align; cell.border = thin_border
            ws.row_dimensions[1].height = 22

        next_row = ws.max_row + 1
        added = skipped = 0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        for job in self.jobs:
            url = job.get("url", "")
            if url in existing_urls:
                skipped += 1
                continue
            existing_urls.add(url)
            fill = odd_fill if (next_row % 2) else even_fill
            desc = job.get("description", "")
            row_data = [
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                job.get("salary_min") or "",
                job.get("salary_max") or "",
                job.get("created", ""),
                url, desc,
                job.get("source", ""),
                now_str,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=val)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 8))
                if col == 7 and url:
                    cell.font = link_font; cell.hyperlink = url
                else:
                    cell.font = data_font
            ws.row_dimensions[next_row].height = 18
            next_row += 1; added += 1

        for col, w in enumerate([40, 25, 22, 16, 16, 14, 40, 60, 16, 18], 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = w

        if "Search Info" in wb.sheetnames:
            del wb["Search Info"]
        meta    = wb.create_sheet("Search Info")
        bg_fill = PatternFill("solid", start_color="1e293b")
        m_font  = Font(name="Arial", size=10, color="F1F5F9")
        m_bold  = Font(name="Arial", size=10, color="38bdf8", bold=True)
        total   = ws.max_row - 1

        for r, (label, value) in enumerate([
            ("Search query",       self.what_var.get()),
            ("Location",           self.where_var.get()),
            ("Country",            self.country_var.get()),
            ("Last updated",       now_str),
            ("Total jobs",         total),
            ("Added this run",     added),
            ("Duplicates skipped", skipped),
            ("Source",             "Adzuna / Reed"),
        ], 1):
            c1 = meta.cell(row=r, column=1, value=label)
            c2 = meta.cell(row=r, column=2, value=value)
            c1.font = m_bold; c1.fill = bg_fill
            c2.font = m_font; c2.fill = bg_fill

        meta.column_dimensions["A"].width = 22
        meta.column_dimensions["B"].width = 35
        meta.sheet_view.showGridLines = False
        ws.sheet_view.showGridLines   = False

        wb.save(path)
        self._set_status(
            f"💾 Saved — {added} new · {skipped} duplicates skipped · {total} total  →  {os.path.basename(path)}")
        self._analytics_populate_files()

        what    = self.what_var.get().strip()
        where   = self.where_var.get().strip()
        country = COUNTRIES.get(self.country_var.get(), "de")
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        if not any(s.get("what") == what and s.get("where") == where for s in searches):
            searches.append({"what": what, "where": where, "country": country})
            _save_config({"searches": searches})
            self._autorun_refresh()

    # ── Schedule Task ─────────────────────────────────────────────────────────
    def _schedule(self):
        if sys.platform != "win32":
            messagebox.showinfo("Not supported", "Task Scheduler is only available on Windows.")
            return

        time_str = self.sched_time_var.get().strip()
        try:
            hour, minute = time_str.split(":")
            int(hour); int(minute)
        except Exception:
            messagebox.showerror("Invalid time", "Please enter time as HH:MM (e.g. 08:00)")
            return

        # Save time and ensure searches list exists
        cfg      = _load_config()
        searches = cfg.get("searches", [])
        if not searches:
            self.sched_status.config(
                text="Add at least one search to the list first.", fg=DANGER)
            return
        _save_config({"sched_time": time_str})

        # When running as a frozen .exe, sys.executable is the .exe itself — find Python from PATH
        if getattr(sys, "frozen", False):
            import shutil
            python_exe = shutil.which("python") or shutil.which("python3")
            if not python_exe:
                messagebox.showerror("Python not found",
                    "Could not find Python in your system PATH.\n\n"
                    "Please install Python and make sure it is added to PATH,\n"
                    "then try again.")
                return
        else:
            python_exe = sys.executable

        script_path = os.path.join(_app_dir(), "job_search_auto.py")
        task_name   = "JobSearchDaily"

        if not os.path.exists(script_path):
            messagebox.showerror("Missing file",
                                 f"job_search_auto.py not found:\n{script_path}\n\n"
                                 "Please place job_search_auto.py in the same folder.")
            return

        ps_script = f"""
$action  = New-ScheduledTaskAction -Execute '"{python_exe}"' -Argument '"{script_path}"'
$trigger = New-ScheduledTaskTrigger -Daily -At '{hour.zfill(2)}:{minute.zfill(2)}'
$settings = New-ScheduledTaskSettingsSet `
    -AllowStartIfOnBatteries `
    -DontStopIfGoingOnBatteries `
    -StartWhenAvailable `
    -ExecutionTimeLimit (New-TimeSpan -Hours 1)
Register-ScheduledTask -TaskName '{task_name}' -Action $action -Trigger $trigger `
    -Settings $settings -Force | Out-Null
"""
        result = subprocess.run(
            ["powershell", "-NonInteractive", "-Command", ps_script],
            capture_output=True, text=True)

        searches = _load_config().get("searches", [])
        if result.returncode == 0:
            t = f"{hour.zfill(2)}:{minute.zfill(2)}"
            self._set_status(f"✓ Scheduled daily at {t}")
            self.sched_status.config(
                text=f"✓ Scheduled daily at {t} — {len(searches)} search(es) active", fg=SUCCESS)
            self._refresh_sched_status()
            messagebox.showinfo("Scheduled",
                f"Daily auto-run set for {t}\n\n"
                f"{len(searches)} search(es) will run every day — even on battery.\n"
                f"Task name: {task_name}\n\n"
                "To remove: click 'Remove Schedule'.")
        else:
            self._set_status("Scheduling failed — try Run as Administrator", ok=False)
            self.sched_status.config(text="Scheduling failed — try Run as Administrator", fg=DANGER)
            messagebox.showerror("Error",
                f"Could not create the scheduled task.\n\n"
                f"{result.stderr.strip()}\n\n"
                "Try right-clicking the app and selecting 'Run as administrator'.")

    # ── Remove Scheduled Task ─────────────────────────────────────────────────
    def _unschedule(self):
        if sys.platform != "win32":
            messagebox.showinfo("Not supported", "Task Scheduler is only available on Windows.")
            return

        task_name = "JobSearchDaily"
        confirm = messagebox.askyesno(
            "Remove Schedule",
            f"Remove the daily auto-run task '{task_name}'?\n\n"
            "The app and Excel files are not affected.")
        if not confirm:
            return

        result = subprocess.run(
            f'schtasks /delete /f /tn "{task_name}"',
            shell=True, capture_output=True, text=True)

        if result.returncode == 0:
            self._set_status("✓ Daily schedule removed")
            self._refresh_sched_status()
            messagebox.showinfo("✓ Removed",
                f"Task '{task_name}' has been removed.\n\n"
                "You can re-schedule anytime with ⏱ Schedule Daily.")
        else:
            err = result.stderr.strip().lower()
            if "cannot find" in err or "nicht gefunden" in err:
                self._set_status("⚠ No active schedule found", ok=False)
                messagebox.showwarning("Not found",
                    f"No task named '{task_name}' was found.\n"
                    "It may have already been removed.")
            else:
                self._set_status("⚠ Could not remove schedule", ok=False)
                messagebox.showerror("Error", f"schtasks error:\n{result.stderr}")


if __name__ == "__main__":
    try:
        JobSearchApp().mainloop()
    except Exception as e:
        _log_error(type(e), e, e.__traceback__)
        traceback.print_exc()
        input("Press Enter to close...")
